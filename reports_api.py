"""
Author 1: Satyanarayana Y
Author 2: Gopi Teja B
Created Date: 18-08-2022
"""

from ace_logger import Logging
from db_utils import DB
from datetime import datetime,timedelta
from flask import request, jsonify
from pathlib import Path
from py_zipkin.util import generate_random_64bit_string
from app import app
#from app.elasticsearch_utils import elasticsearch_search
from py_zipkin.zipkin import zipkin_span,ZipkinAttrs, create_http_headers_for_new_span
from elasticsearch import Elasticsearch
import ast
import openpyxl
# import xlsxwriter
from openpyxl import load_workbook

try:
    from generate_reports import reports_consumer, generate_report
except Exception as e:
    from app.generate_reports import reports_consumer, generate_report
from time import time as tt
from pandas import Timestamp
from openpyxl.styles import NamedStyle

import base64
import json
import uuid
import pytz
import requests
import os
import re
import datetime as dt
import psutil
import pandas as pd
import numpy as np



es_dns = os.environ.get('ELASTIC_SEARCH_FULL_SEARCH_DNS','')
es_port = os.environ.get('ELASTIC_SEARCH_FULL_PORT', '')
es_scheme = os.environ.get('ELASTIC_SEARCH_FULL_SEARCH_SCHEME','')
es = Elasticsearch(
    [f'{es_dns}'],
    http_auth=('elastic','MagicWord'),
    scheme=f"{es_scheme}",
    port=es_port,
)


logging = Logging(name='reports_api')

db_config = {
    'host': os.environ['HOST_IP'],
    'password': os.environ['LOCAL_DB_PASSWORD'],
    'user': os.environ['LOCAL_DB_USER'],
    'port': os.environ['LOCAL_DB_PORT']
}


def measure_memory_usage():
    process = psutil.Process()
    memory_info = process.memory_info()
    return memory_info.rss  # Resident Set Size (RSS) in bytes


def http_transport(encoded_span):

    body = encoded_span
    requests.post(
        'http://servicebridge:5002/zipkin',
        data=body,
        headers={'Content-Type': 'application/x-thrift'},
    )


def get_reports_column_mapping(columns):
    return_columns_map = {}
    if columns:
        for column in columns:
            return_columns_map[column.replace('_', ' ').title()] = column
    return return_columns_map

def insert_into_audit(data):
    tenant_id = data.pop('tenant_id')
    db_config['tenant_id'] = tenant_id
    stats_db = DB('stats', **db_config)
    stats_db.insert_dict(data, 'audit')
    return True

db_column_types = {
    "mysql": {
        "int": "number",
        "tinyint": "number",
        "smallint": "number",
        "mediumint": "number",
        "bigint": "number",
        "double": "number",

        "date": "date",
        "datetime": "date",
        "timestamp": "date",
        "time": "date",

        "char": "string",
        "varchar": "string",
        "blob": "string",
        "text": "string",
        "tinytext": "string",
        "mediumtext": "string",
        "longtext": "string",
        "tinyblob": "string",
        "mediumblob": "string",
        "longblob": "string",
        "enum": "string"
    },
    "mssql": {
        "bigint": "number",
        "bit": "number",
        "decimal": "number",
        "int": "number",
        "money": "number",
        "numeric": "number",
        "smallint": "number",
        "smallmoney": "number",
        "tinyint": "number",
        "float": "number",

        "date": "date",
        "datetime2": "date",
        "datetime": "date",
        "datetimeoffset": "date",
        "smalldatetime": "date",
        "time": "date",

        "char": "string",
        "text": "string",
        "varchar": "string",
        "nchar": "string",
        "nvarchar": "string",
        "ntext": "string",


    },
    "oracle": {
        "NUMBER": "number",
        "BINARY_FLOAT": "number",
        "BINARY_DOUBLE": "number",
        "DATE": "date",
        "TIMESTAMP": "date",
        "CHAR": "string",
        "NCHAR": "string",
        "VARCHAR2": "string",
        "NVARCHAR2": "string",
        "CLOB": "string",
        "NCLOB": "string",
        "BLOB": "string"
    }
}


def create_index(tenant_ids, sources=[]):
    body = {
        "settings": {
            "analysis": {
                "analyzer": {
                    "default": {
                        "type": "custom",
                        "tokenizer": "whitespace",
                        "filter": [
                            "lowercase"
                        ]
                    }
                }
            }
        },
        "mappings": {
            "date_detection": "false",
            "numeric_detection": "false"
        }
    }
    
    body_with_date = {
            "settings": {
            "analysis": {
                "analyzer": {
                    "default": {
                        "type": "custom",
                        "tokenizer": "whitespace",
                        "filter": [
                          "lowercase"
                        ]
                    }
                }
            }
            },
        "mappings": {
            "properties": {
                "@timestamp": {
                    "type": "date"
                },
                "@version": {
                    "type": "text",
                    "fields": {
                        "keyword": {
                            "type": "keyword",
                            "ignore_above": 256
                        }
                    }
                },
                "ocr": {
                    "properties": {
                        "created_date": {
                            "type": "date"
                        },
                        "last_updated": {
                            "type": "date"
                        }
                    }
                },
                "process_queue": {
                    "properties": {
                        "created_date": {
                            "type": "date"
                        },
                        "last_updated": {
                            "type": "date"
                        },
                        "freeze": {
                            "type": "boolean"
                        },
                        "case_lock":{
                            "type": "boolean"
                        }
                    }
                }
            },
            "date_detection": "false",
            "numeric_detection": "false"
        }
        }
    indexes = []
    
    for tenant_id in tenant_ids:
        indexes.extend(get_search_indexes(sources, tenant_id))

    for ind in indexes:
        es.indices.delete(index=ind, ignore=[400, 404])
        if 'processqueue' in ind:
            es.indices.create(index=ind, body=body_with_date, ignore=400)
        else:
            es.indices.create(index=ind, body=body, ignore=400)


def get_search_indexes(sources, temp_tenant_id=''):
    """
    Author : Akshat Goyal
    :param sources:
    :return:
    """
    tenant_id = temp_tenant_id.replace('.', '')
    if not sources:
        return '_all'
    indexes = []
    if isinstance(sources, list):
        for source in sources:
            new_source = tenant_id + source if tenant_id else source
            new_source = new_source.replace('.', '').replace('_', '')
            indexes.append(new_source)
    elif isinstance(sources, str):
        new_source = tenant_id + '_' + sources if tenant_id else sources
        new_source = new_source.replace('.', '').replace('_', '')
        indexes.append(new_source)

    return indexes


def master_search(tenant_id, text, table_name, start_point, offset, columns_list, header_name):
    elastic_input = {}
    
    print(f'#######{tenant_id}')
    print(f'#######{table_name}')
    print(f'#######{start_point}')
    print(f'#######{offset}')
    print(f'#######{columns_list}')
    print(f'#######{header_name}')
    elastic_input['columns'] = columns_list
    elastic_input['start_point'] = start_point
    elastic_input['size'] = offset
    if header_name:
        header_name=header_name.upper()
        elastic_input['filter'] = [{'field': header_name, 'value': "*" + text + "*"}]
    else:
        elastic_input['text'] = text
    elastic_input['source'] = table_name
    elastic_input['tenant_id'] = tenant_id
    print(f"output of the elastic_input---------{elastic_input}")
    files, total = elasticsearch_search(elastic_input)
    print(f"output----------{files,total}")
    return files, total


def get_reports_column_data_types(columns, tenant_id):

    logging.info(f"^^^^^^^^^^^^^^^ COLUMNS ARE {columns}")

    db_config['tenant_id'] = tenant_id
    reports_db = DB('reports', **db_config)

    column_query = """SELECT *
        FROM report_requests
        FETCH FIRST 1 ROW ONLY"""
    reports_columns = list(reports_db.execute_(column_query).columns.values)


    query = """SELECT DATA_TYPE
        FROM USER_TAB_COLUMNS
        WHERE TABLE_NAME = 'REPORT_REQUESTS'"""
    reports_column_types = list(reports_db.execute_(query)['data_type'])

    reports_dt_map = dict(zip(reports_columns, reports_column_types))

    logging.info(
        f"#########################33 reports VALUE IS {reports_dt_map} ")

    col_data_types = {}

    db_type = os.environ.get('DB_TYPE', 'mysql').lower()

    for column in columns:
        col_data_types[column] = db_column_types[db_type].get(
            reports_dt_map.get(column, ""), "unknown")

    return col_data_types



def get_group_ids(user, db):
    logging.info(f'Getting group IDs for user `{user}`')


    query="""SELECT organisation_attributes.attribute, user_organisation_mapping.value 
        FROM user_organisation_mapping 
        JOIN active_directory ON active_directory.id = user_organisation_mapping.user_id 
        JOIN organisation_attributes ON organisation_attributes.id = user_organisation_mapping.organisation_attribute 
        WHERE active_directory.username = %s"""

    user_group = db.execute_(query, params=[user])

    if user_group.empty:
        logging.error(f'No user organisation mapping for user `{user}`')
        return

    user_group_dict = dict(zip(user_group.attribute, user_group.value))
    user_group_dict = {key: [value] for key, value in user_group_dict.items()}
    query="select * from GROUP_DEFINITION"
    group_def_df = db.execute_(query)

    if group_def_df.empty:
        logging.debug('Groups not defined in `group_definition`')

        return
    logging.info("####groupgroup_def_df {group_def_df}")

    user_group = group_def_df.to_dict()
    group_def = {}
    for i in range(len(user_group['id'])):
        id_key = str(user_group['id'][i])
        group_def[id_key] = {
            'group_name': user_group['group_name'][i],
            'group_definition': user_group['group_definition'][i],
            'group_definition_template': user_group['group_definition_template'][i]
        }

    logging.info(f"####ggroup_def {group_def}")
    group_ids = []
    for index, group in group_def.items():
        logging.debug(f'Index: {index}')
        logging.debug(f'Group: {group}')

        try:
            group_dict = json.loads(group['group_definition'])
        except Exception:
            logging.error('Could not load group definition dict.')
            break


        if group_dict.items() == user_group_dict.items():
            group_ids.append(index)

    logging.info(f'Group IDs: {group_ids}')
    return group_ids

@app.route('/get_reports_queue', methods=['POST', 'GET'])
def get_reports_queue():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("##Failed to start ram and time calc")
        
    data = request.json

    logging.info(f'Data recieved: {data}')
    tenant_id = data.get('tenant_id', None)
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='get_reports_queue',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):
        try:
            
            user = data.get('user', None)
            total_count = 0  
            filtered_text = data.get('filter_text','')
            try:
                start_point = data.get('start', 1)
                end_point = data.get('end', 20)
            except Exception as e:
                logging.debug(e)
                start_point = 1
                end_point = 20

            try:
                offset = end_point - start_point
            except Exception:
                start_point = 1
                end_point = 20
                offset = 20
            offset_ = 0 if start_point == 1 else start_point - 1

            # Sanity check
            if user is None:
                message = 'User not provided in request.'
                logging.error(message)
                return jsonify({'flag': False, 'message': message})

            # Connect to reports database
            reports_db_config = {
                                'host': os.environ['HOST_IP'],
                                'password': os.environ['LOCAL_DB_PASSWORD'],
                                'user': os.environ['LOCAL_DB_USER'],
                                'port': os.environ['LOCAL_DB_PORT'],
                                'tenant_id': tenant_id
                            }
            reports_db = DB('reports', **reports_db_config)


            user__=user

            # Fetch reports
            logging.debug(f'########### Fetching reports for user `{user}`')


            if filtered_text!='':
                text = filtered_text
                table_name = 'report_requests'
                start_point = 0
                end_point = 9999
                header_name = None
                offset = end_point - start_point
                columns_list = ['REFERENCE_ID', 'REPORT_NAME', 'GENERATED_DATETIME', 'REQUESTED_BY', 'STATUS', 'REQUESTED_DATETIME','TAGS']
                columns_list_ = ["Reference Id","Report Name","Generated Datetime","Requested By","Status","Requested Datetime","Tags"]
                files, total = master_search(tenant_id = tenant_id, text = text, table_name = table_name, start_point = 0, offset = 9999, columns_list = columns_list, header_name=header_name)

                if end_point > total:
                    end_point = total
                if start_point == 1:
                    logging.info("start_point is 1")                
                else:
                    start_point += 1
                print(f'#########Files got are: {files}')
                data = pd.DataFrame(files)
                filtered_df = data[data['REQUESTED_BY'] == user]
                files = filtered_df.to_dict(orient='records')
                print(f'Files got are: {files}')
                for i in files:
                    for key,value in i.items():
                        if key=='TAGS' or key=='tags':
                            try:
                                i[key]=ast.literal_eval(value)
                            except Exception:
                                pass
                end_point = len(files)
                column_data_types = {"GENERATED_DATETIME":"unknown","REFERENCE_ID":"unknown","REPORT_NAME":"unknown","REQUESTED_BY":"unknown","REQUESTED_DATETIME":"unknown","STATUS":"unknown","TAGS":"unknown","generated_datetime":"string","reference_id":"number","report_name":"string","requested_by":"string","requested_datetime":"string","status":"date","tags":"string"}
                column_mapping = {"Generated Datetime":"GENERATED_DATETIME","Reference Id":"REFERENCE_ID","Report Name":"REPORT_NAME","Requested By":"REQUESTED_BY","Requested Datetime":"REQUESTED_DATETIME","Status":"STATUS","Tags":"TAGS"}
                pagination = {"start": start_point, "end": end_point, "total": end_point}
                final_data = []
                for i in files:
                    dic={}
                    for key,value in i.items():
                        key=key.lower()
                        dic[key]=value
                    res_dic = {**i,**dic}
                    final_data.append(res_dic)
                respose_data = {
                'flag': True,
                'data': {
                    'files': final_data,
                    'buttons': [],
                    'field': [],
                    'tabs': [],
                    'column_data_types': column_data_types,
                    'column_mapping': column_mapping,
                    'excel_source_data': {},
                    'tab_type_mapping': {},
                    'pagination': pagination,
                    'column_order': columns_list_,
                    'children_dropdown': [],
                    'biz_rules': [],
                    'dropdown_values': {},
                    'cascade_object': "{}",
                    'get_report_view' : True}
                }
                return jsonify(response_data)

            pagination_query="""
            SELECT REFERENCE_ID, REPORT_NAME, GENERATED_DATETIME, REQUESTED_BY, STATUS, REQUESTED_DATETIME,TAGS
            FROM report_requests 
            WHERE requested_by = %s AND parent_id IS NULL 
            ORDER BY requested_datetime DESC 
            OFFSET %s ROWS FETCH NEXT %s ROWS ONLY"""
            try:
                user_reports_data = reports_db.execute_(
                    pagination_query, params=[user,  offset_ , offset+1])

                user_reports_data_json = []

                logging.info(
                    f"############### user_reports_data {user_reports_data}")

                for idx, value in enumerate(user_reports_data["generated_datetime"].isnull()):
                    user_reports_data.loc[idx, "requested_datetime"] = str(
                        user_reports_data.loc[idx, "requested_datetime"])
                    if value:
                        user_reports_data.loc[idx, "generated_datetime"] = None
                    else:
                        user_reports_data.loc[idx, "generated_datetime"] = str(
                            user_reports_data.loc[idx, "generated_datetime"])

                user_reports_data_json = user_reports_data.to_dict('records')
                for i, report_data in enumerate(user_reports_data_json):
                    logging.info(f"############################ {report_data}")
                    if report_data['tags'] == "" or report_data['tags'] == '' or report_data['tags'] == None:
                        report_data['tags'] = {}
                    else:
                        report_data['tags'] = json.loads(report_data['tags'])
                    try:
                        report_data['requested_by']=user__
                        #report_data['generated_datetime'] = datetime.strptime(report_data['generated_datetime'], "%Y-%m-%d %H:%M:%S")
                        #report_data['generated_datetime']=report_data['generated_datetime'].strftime('%d-%b-%Y %H:%M:%S')
                        #report_data['requested_datetime']=datetime.strptime(report_data['requested_datetime'], "%Y-%m-%d %H:%M:%S")
                        #report_data['requested_datetime']=report_data['requested_datetime'].strftime('%d-%b-%Y %H:%M:%S')
                        if isinstance(report_data['generated_datetime'], datetime):
                            report_data['generated_datetime'] = report_data['generated_datetime'].strftime('%d-%b-%Y %H:%M:%S')
                        else:
                            report_data['generated_datetime'] = datetime.strptime(report_data['generated_datetime'], "%Y-%m-%d %H:%M:%S")
                            report_data['generated_datetime'] = report_data['generated_datetime'].strftime('%d-%b-%Y %H:%M:%S')

                        if isinstance(report_data['requested_datetime'], datetime):
                            report_data['requested_datetime'] = report_data['requested_datetime'].strftime('%d-%b-%Y %H:%M:%S')
                        else:
                            report_data['requested_datetime'] = datetime.strptime(report_data['requested_datetime'], "%Y-%m-%d %H:%M:%S")
                            report_data['requested_datetime'] = report_data['requested_datetime'].strftime('%d-%b-%Y %H:%M:%S')

                    except Exception as e:
                        print(e)
                        
                    print(f'report_data##{report_data}')
                    try:
                        report_data['TAGS']=json.loads(report_data['TAGS'])
                    except Exception:
                        pass
                 

                    user_reports_data_json[i] = report_data

            except Exception:
                message = 'Error fetching data from database'
                logging.exception(message)
                return jsonify({'flag': False, 'message': message})

            total_reports_query = f"SELECT COUNT(*) AS COUNT FROM report_requests where requested_by='{user}'"
            total_reports_df = reports_db.execute_(total_reports_query)

            total_count = list(total_reports_df['COUNT'])[0]
            logging.debug(total_count)



            # Get group_id to know which reports are accessable
            group_db = DB("group_access", **reports_db_config)
            user_groups = get_group_ids(user, group_db)
            logging.info(f"##### user_groups: {user_groups}")

            active_directory_user_query = f"SELECT ROLE FROM `active_directory` where username = '{user}'"
            active_directory_user_df = group_db.execute_(active_directory_user_query)
            user_role = active_directory_user_df.iloc[0]['ROLE']

            uam_report_ids=[81,105,106,142]
            operation_report_ids=[161,41,42,43,61,181,201,221]
            reports_query_data=[]
            query=f"select display_role_rights,new_rights_assigned_status from role_rights where display_role_rights in ('UAM Reports','Operation Reports') and role_name='{user_role}'"
            reports_right_status=group_db.execute_(query).to_dict(orient= 'records')
            for record in reports_right_status:
                if (record["display_role_rights"] == "UAM Reports" and record["new_rights_assigned_status"].lower() == "yes"):
                    reports_query_data.extend(uam_report_ids)
                if (record["display_role_rights"] == "Operation Reports" and record["new_rights_assigned_status"].lower() == "yes"):
                    reports_query_data.extend(operation_report_ids)
            # reports_query = f"select reports_id from reports_access where group_id = {user_groups[0]}"
            # reports_query_data = group_db.execute_(reports_query)['reports_id'].tolist()

            if len(reports_query_data) > 1:
                reports_query_data = tuple(reports_query_data)
                reports_query_data = "in " + str(reports_query_data)
            else:
                reports_query_data = "= " + str(reports_query_data[0])

            published_flag = 1
            reports_query = f"SELECT report_id,report_name FROM report_master where publish_flag = {published_flag} and parent_id is NULL and report_id {reports_query_data}"
            report_df = reports_db.execute_(reports_query)

            report_dict = []

            report_dict = report_df.to_dict('records')
            logging.debug(f"############ reports dict {report_dict}")

            for i, report in enumerate(report_dict):
                filters_query = f"select * from report_filter where report_id={report['report_id']}"
                filter_list = reports_db.execute_(filters_query).to_dict('records')
                if len(filter_list) > 0:
                    for j, filter_data in enumerate(filter_list):
                        if filter_data['filter_options'] != "" and filter_data['filter_options'] is not None:
                            string_list = filter_data["filter_options"].split("#$")
                            logging.info(
                                f"########### FIlter : {filter_data}and Filter Options:{string_list}")
                            if len(string_list) > 0:
                                for k, string_json in enumerate(string_list):
                                    if string_json != '' and string_json != "":
                                        string_list[k] = json.loads(string_json)
                            filter_list[j]["filter_options"] = string_list
                report['filters'] = filter_list
                report_dict[i] = report

            logging.debug(f"start {start_point} end {end_point} offset {offset}")

            logging.debug('***')
            logging.debug(f"start {start_point} end {end_point} offset {offset}")
            if start_point > end_point:
                start_point = end_point
            logging.debug('&&&&')
            logging.debug(f"start {start_point} end {end_point} offset {offset}")
            if end_point > total_count:
                end_point = total_count
            else:
                end_point = (int(offset) + int(start_point))
            logging.debug('####')
            logging.debug(f"start {start_point} end {end_point} offset {offset}")

            logging.debug(f"end : {end_point}")

            reports_columns = [col_name for col_name in user_reports_data.columns if col_name not in [
                'rn', 'report_id', 'request_id']]

            reports_coulmn_mapping = get_reports_column_mapping(reports_columns)
            reports_column_order = [] if not reports_coulmn_mapping else list(
                reports_coulmn_mapping.keys())

            reports_column_dt_types = get_reports_column_data_types(
                reports_columns, tenant_id)
            print(user_reports_data_json)
            respose_data = {
                'flag': True,
                'data': {
                    'files': user_reports_data_json,
                    'buttons': [],
                    'field': [],
                    'tabs': [],
                    'excel_source_data': {},
                    'tab_type_mapping': {},
                    'pagination': {
                        'start': start_point,
                        'end': end_point,
                        'total': total_count},

                    'column_mapping': reports_coulmn_mapping,
                    'column_order': reports_column_order,
                    'children_dropdown': [],
                    'pdf_type': 'folder' if tenant_id else 'blob',
                    'biz_rules': [],
                    'dropdown_values': {},
                    'cascade_object': "{}",
                    'column_data_types': reports_column_dt_types,
                    'reports': report_dict,
                    'get_report_view' : True
                }
            }

            
        except Exception:
            logging.exception(
                'Something went wrong while getting reports queue. Check trace.')
            response_data = {
                'flag': False, 'message': 'System error [/get_reports_queue]! Please contact your system administrator.'}
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except Exception:
            logging.warning("#Failed to calc end of ram and time")
            logging.exception("#ram calc went wrong@@@")
            memory_consumed = None
            time_consumed = None
            
        logging.info(f"## Reports API get_reports_queue Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")

        return jsonify(respose_data)
    

@app.route('/generate_reports', methods=['POST', 'GET'])
def generate_reports():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("###Failed to start ram and time calc")
    # Get data from UI
    data = request.json
    logging.info(f"## Reports info request data -- {data}")
    tenant_id = data.get('tenant_id', None)
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='generate_reports',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):
        
        requested_by = data.get('user', None)
        report_name = data.get('report_name', None)
        report_id = data.get("report_id", -1)

        filters = data.get('filters', {})
        try:
            filters["start_date"]=filters["start_date"]+' 00:00:00'
            filters["end_date"]=filters["end_date"]+' 23:59:59'
        except Exception:
            filters["start_date"]='00-00-00 00:00:00'
            filters["end_date"]='00-00-00 00:00:00'


        text_filters = json.dumps(filters)

        # Sanity check
        if tenant_id is None:
            message = 'tenant_id not provided in request.'
            logging.error(message)
            return jsonify({'flag': False, 'message': message})

        if report_name is None:
            message = 'Report type not provided in request.'
            logging.error(message)
            return jsonify({'flag': False, 'message': message})

        logging.info(f"debugging report_name from front_end : {report_name}")
        logging.info(f"debugging report_id from front_end : {report_id}")

        # Generate file name if not given
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        timestamp_actual = timestamp
        timestamp1 = timestamp.strftime('%d-%m-%Y %H:%M:%S')

        timestamp = str(timestamp)[:-13]
        # Add file to database
        reports_db_config = {
                                'host': os.environ['HOST_IP'],
                                'password': os.environ['LOCAL_DB_PASSWORD'],
                                'user': os.environ['LOCAL_DB_USER'],
                                'port': os.environ['LOCAL_DB_PORT'],
                                'tenant_id': tenant_id
                            }
        logging.debug(reports_db_config)


        # FETCHING DATA FROM REPORT CONFIG TABLES
        report_master_db = DB('reports', **reports_db_config)

        sheet_name = None
        report_id_idx = 0
        parent_start_time = -1
        parent_ref_id = -1
        parent_req_id = -1
        parent_id = -1
        parent_query = f"SELECT report_id FROM report_master where parent_id = '{report_id}'"
        parent_id_val = report_master_db.execute_(parent_query)
        report_ids = list(parent_id_val['report_id'])
        if len(report_ids) == 0:
            report_ids = [report_id]
        else:
            parent_id = report_id
            report_ids.insert(0, report_id)

        logging.debug(f"List of report ids to process {report_ids}")
        for report_id in report_ids:
            logging.info(f"processing this {report_id}")
            # Generate a reference ID (10 digits)
            reference_id = uuid.uuid4().hex[:10].upper()
            if report_id_idx == 0:
                parent_ref_id = reference_id
            query = f'SELECT report_filename,report_out_format FROM report_template WHERE report_id={report_id}'
            report_data = report_master_db.execute_(query).to_dict('records')[0]

            filename_var = report_data["report_filename"]

            # If the file name in report Template is not mentioned it will  take report name as file name
            if filename_var == "" or filename_var == None:
                filename_var = report_name

            # Checking Fund Name
            fund_name_var = ""
            if 'fund_name' in filters:
                fund_name_var = filters['fund_name']

            file_type = report_data["report_out_format"]
            if parent_id == -1:
                file_name = f'{filename_var}-{fund_name_var}-{timestamp1}.{file_type}'
            else:
                if parent_ref_id == reference_id:
                    file_name = ""
                else:
                    file_name = f'{filename_var}-{timestamp1}#{reference_id}.{file_type}'
                    report_sheet_name_query = f"select report_name from report_master where report_id={report_id}"
                    sheet_name = report_master_db.execute_(
                        report_sheet_name_query).values.tolist()[0][0]
                    report_name = sheet_name
                    if report_id_idx == 0:
                        parent_start_time = 1

            base_query = f"select query_type,report_query,route from report_master where report_id={report_id}"

            base_query_df = report_master_db.execute_(base_query)
            base_query_list = base_query_df['report_query'].to_list()
            route_list = base_query_df['route'].tolist()
            query_type = base_query_df['query_type'].to_list()[0]

            # Updating ETA for reports
            average_query = f"SELECT AVG(process_time) AS ptime FROM report_requests WHERE report_id={report_id} AND report_name   ='{report_name}' AND status='Download' AND process_time IS NOT NULL"
            report_eta = report_master_db.execute_(average_query)
            eta = 0
            if len(report_eta['ptime']) > 0 and report_eta['ptime'][0] != None:
                eta = int(report_eta['ptime'][0])
           
            formatted_requested_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            query = f"SELECT SUM(process_time) AS sum FROM report_requests WHERE status = 'Processing' AND requested_datetime < TO_DATE('{formatted_requested_time}', 'YYYY-MM-DD HH24:MI:SS')"
            report_eta = report_master_db.execute_(query)
            processing = 0
            if len(report_eta['sum']) > 0 and report_eta['sum'][0] != None:
                processing = int(report_eta['sum'][0])
            eta += processing
            eta_datetime = dt.timedelta(seconds=int(eta))
            eta_datetime += timestamp_actual

            eta_datetime = str(eta_datetime)[:-13]

            filters_data = ""
            exec_query = ""
            if len(base_query_list) > 0 and query_type == 'query':
                exec_query = base_query_list[0]
                if bool(filters):


                    for column, value in filters.items():

                        if "{"+column in exec_query:
                            if len(report_ids) == 1:
                                column_query = f"select * from report_filter where report_id={report_id} and unique_name='{column}'"
                            elif len(report_ids) > 1:
                                column_query = f"select * from report_filter where report_id={parent_id} and unique_name='{column}'"

                            try:
                                column_df = report_master_db.execute_(column_query)
                                filter_type = column_df['filter_type'].to_list()[0]
                                logging.info("########## FILTER TYPE: ", filter_type)

                                if filter_type.strip().lower() == "string":
                                    exec_query = exec_query.replace(
                                        "{"+column+"}", "'"+value+"'")
                                    

                                elif filter_type.strip().lower() == "int" or filter_type.strip().lower() == "integer":
                                    exec_query = exec_query.replace(
                                        "{"+column+"}", value)
                                    
                                elif filter_type.strip().lower() == "date_picker":
                                    logging.debug(f"BEFORE {exec_query}")
                                    exec_query = exec_query.replace(
                                        "{"+column+"}", "'"+value+"'")
                                    logging.debug(f"AFTER {exec_query}")
                                elif filter_type.strip().lower() == "date_range":
                                    if "{"+column+"_begin"+"}" in exec_query:
                                        exec_query = exec_query.replace(
                                            "{"+column+"_begin"+"}", "'"+value['begin']+"'")
                                    elif "{"+column+"_start"+"}" in exec_query:
                                        exec_query = exec_query.replace(
                                            "{"+column+"_start"+"}", "'"+value['begin']+"'")
                                    if "{"+column+"_end"+"}" in exec_query:
                                        exec_query = exec_query.replace(
                                            "{"+column+"_end"+"}", "'"+value['end']+"'")
                                elif filter_type.strip().lower() == "float":
                                    logging.info("filter_type is float")

                            except Exception as e:
                                logging.exception(e)
                                return jsonify({'flag': False, 'message': f'Failed to get the filter data {column}'})

            elif len(route_list) > 0 and query_type == 'route':
                if len(filters) > 0:
                    filters_data = json.dumps(filters)
                elif len(filters) == 0:
                    filters_data = json.dumps({})
                exec_query = route_list[0]

            if len(route_list) > 0 and report_name == 'Audit_Report':
                print(f"entered report name is Audit........................")
                if len(filters) > 0:
                    filters_data = json.dumps(filters)
                elif len(filters) == 0:
                    filters_data = json.dumps({})
                #print(f"exec query is after :{exec_query}")

            report_id_idx = report_id_idx + 1
            reports_db = DB('reports', **reports_db_config)
            file_name = str(file_name).replace(' ', '_')
            file_name = str(file_name).replace(':', '_')
            logging.debug(f"################### File name after {file_name}")
            formatted_requested_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            formatted_eta_datetime = datetime.strptime(eta_datetime, '%Y-%m-%d %H:%M:%S')
            # print(formatted_requested_time,type(formatted_requested_time),'###req')
            # print(formatted_eta_datetime,type(formatted_eta_datetime),"##eta")
            # logging.info(type(eta_datetime),type(timestamp),eta_datetime,timestamp)
            insert_data={
                'REFERENCE_ID': reference_id,
                'REPORT_NAME': report_name,
                'REQUESTED_BY': requested_by,
                'REPORT_OUTPUT': file_name,
                'STATUS': 'Processing',
                'REQUESTED_DATETIME': formatted_requested_time,
                'ACTUAL_REPORT_QUERY': exec_query,
                'REPORT_ID': str(report_id),
                'QUERY_TYPE': query_type,
                'TAGS': filters_data,
                'QUERY_PARAMS': filters_data,
                'FILTERS': text_filters,
                'FUND_NAME': fund_name_var,
                'ETA': formatted_eta_datetime
            }

            # logging.info(f"##################### INSERT DATA {insert_data}")
            report_master_db.insert_dict(insert_data, 'report_requests')
            logging.info(f"debug sheet name : {sheet_name}")
            logging.info(f"debug report name : {report_name}")
            logging.info(f"debug parent_id : {parent_id}")
            logging.info(f"debug report_id : {report_id}")
            if parent_id != -1:
                if parent_ref_id == reference_id:
                    logging.info("################Parent request id generated")
                    parent_req_id_query = "SELECT MAX(request_id) FROM report_requests"
                    parent_req_id = report_master_db.execute_(
                        parent_req_id_query).values[0][0]

                if parent_ref_id != reference_id:
                    query = f'SELECT `tags` from report_requests where reference_id = "{reference_id}"'
                    tags_df = report_master_db.execute_(query)
                    tags_query_list = tags_df['tags'].to_list()
                    tags_ = tags_query_list[0]
                    query_ex = f"UPDATE report_requests SET tags='{tags_}' WHERE reference_id = '{parent_ref_id}'"
                    report_master_db.execute_(query_ex)
                    req_id_query = "SELECT MAX(request_id) FROM report_requests"
                    req_id = report_master_db.execute_(req_id_query).values[0][0]
                    query = f"UPDATE report_requests SET parent_id='{parent_req_id}' WHERE request_id={req_id}"
                    template_query = f"UPDATE report_template SET report_sheetname='{sheet_name}' WHERE report_id={report_id}"
                    try:
                        report_master_db.execute_(query)
                        logging.info("sheet name updated in report template")
                    except Exception:
                        logging.info(
                            "Sheet name is not updated in report template")
                    logging.info("before sheetname update")
                    report_master_db.execute_(template_query)
                    logging.info("parent id is updated in report_requests table")

            logging.info('##### sending to generate report')
            if parent_id == -1:
                parent_ref_id = -1
            # Additional data to producer for parent
            insert_data['parent_id'] = str(parent_id)
            insert_data['parent_req_id'] = str(parent_req_id)
            insert_data['parent_ref_id'] = str(parent_ref_id)
            insert_data['parent_start_time'] = str(parent_start_time)
            insert_data['fund_name'] = fund_name_var

            logging.debug(insert_data)
            # Produce to reports consumer
            if report_id != parent_id:
                logging.info(f"########################producing {insert_data}")                               
                result_message= reports_consumer({'tenant_id': tenant_id, 'report_name': report_name, **insert_data})
                logging.info(f"########################result_message {result_message}")
            if parent_id != -1:
                reference_id = parent_ref_id
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except Exception:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("#ram calc went wrong")
            memory_consumed = None
            time_consumed = None
        
        logging.info(f"## Reports API generate_reports Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")
        report_status_query = f"SELECT `status` FROM `report_requests` where reference_id like '%{reference_id}%'"
        report_status = reports_db.execute_(report_status_query).to_dict(orient="records")[0]['status']
        if report_status == 'Failed':
            return jsonify({'flag': True, 'message': f'The report not generated (Ref No. {reference_id})', 'reference_id': reference_id})
        else:
            return jsonify({'flag': True, 'message': f'The report will be available soon to download. (Ref No. {reference_id})', 'reference_id': reference_id})


@app.route('/get_report_view', methods=['POST', 'GET'])
def get_report_view():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("####Failed to start ram and time calc")
        
    data = request.json
    
    tenant_id = data.get('tenant_id', '')

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='get_report_view',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        reference_id = data.get('reference_id',None)
        db_config['tenant_id'] = tenant_id
        db = DB('reports', **db_config)
        query = 'SELECT * FROM `report_requests` WHERE `reference_id`=%s'
        try:
            report_data = db.execute_(query, params=[reference_id]).to_dict('records')[0]
            logging.info(f"############### Report Data: {report_data}")
            html_out = report_data["html_report"]
            if html_out is None:
                request_id = report_data['request_id']
                query1 = f"select * from `report_requests` WHERE parent_id = {request_id}"
                html_out = db.execute_(query1)['html_report'].to_list()
            return_data = {'flag': True, 'data': html_out}
        except Exception:
            logging.error(e)
            message = 'Report preview failed'
            return_data = {'flag': False, 'data': message}
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except Exception as e:
            logging.warning("##Failed to calc end of ram and time")
            logging.exception("###ram calc went wrong##")
            memory_consumed = None
            time_consumed = None
        logging.info(f"## Reports API get_report_view Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")

        return return_data
@app.route('/download_report', methods=['POST', 'GET'])
def download_report():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("#####Failed to start ram and time calc")
    # Get data from UI
    data = request.json
    logging.info(f'Recieved data: {data}')
    tenant_id = data.get('tenant_id', None)
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='download_report',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        reference_id = data.get('reference_id', None)
        report_name = data.get('report_name', None)
        generated_datetime = data.get('generated_datetime','')

        if reference_id is None:
            message = 'Reference ID is not provided.'
            logging.error(message)
            return jsonify({'flag': False, 'message': message})

        # Add file to database
        reports_db_config = {
                                'host': os.environ['HOST_IP'],
                                'password': os.environ['LOCAL_DB_PASSWORD'],
                                'user': os.environ['LOCAL_DB_USER'],
                                'port': os.environ['LOCAL_DB_PORT'],
                                'tenant_id': tenant_id
                            }
        reports_db = DB('reports', **reports_db_config)

        query = 'SELECT * FROM `report_requests` WHERE `reference_id`=%s'
        report_info = reports_db.execute_(query, params=[reference_id])

        report_file_name = list(report_info['report_output'])[0]
        logging.info(f"report_name####{report_name}")
        if 'Completed' in report_name :
            file='completed'
        if 'Rejected' in report_name :
            file='rejected'
        if 'Onhold' in report_name :
            file='onhold'

        if 'Completed' in report_name or 'process report' in report_name or 'Onhold' in report_name or  'Rejected' in report_name :
            #parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            try:
                # Format with weekday + timezone
                parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            except ValueError:
                # Format without weekday + timezone
                parsed_date = datetime.strptime(generated_datetime, "%d-%b-%Y %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'{file}_report--{formatted_date}.xlsx'
        if 'Process Report' in report_name:
            #parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            parsed_date = datetime.strptime(generated_datetime, "%d-%b-%Y %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'Process_Report--{formatted_date}.xlsx'
        
        if 'Extraction Report' in report_name:
            #parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            try:
                # Format with weekday + timezone
                parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            except ValueError:
                # Format without weekday + timezone
                parsed_date = datetime.strptime(generated_datetime, "%d-%b-%Y %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'Extraction_report--{formatted_date}.xlsx'

        if 'Reconciliation Report' in report_name:
            #parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            parsed_date = datetime.strptime(generated_datetime, "%d-%b-%Y %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'Reconciliation_Report--{formatted_date}.xlsx'

        logging.info(f"report_file_name:{report_file_name}")
        report_path = Path(f'./reports/{report_file_name}')
        logging.info(f"@@@@@@@report_path: {report_path}")

        if 'Completed' in str(report_file_name) or 'Process_Report' in str(report_file_name) or 'Onhold' in str(report_file_name) or  'Rejected' in str(report_file_name):
            try:
                with open(report_path, 'rb') as f:
                    report_blob = base64.b64encode(f.read())
            except Exception:
                timestamp_match = re.search(r'_(\d{2})_(\d{2})\.xlsx', str(report_file_name))
                if timestamp_match:
                    hours, minutes = map(int, timestamp_match.groups())
                    timestamp = datetime.strptime(f"{hours:02d}_{minutes:02d}", '%H_%M')
                    updated_timestamp = timestamp - timedelta(minutes=1)
                    updated_timestamp_str = updated_timestamp.strftime('%H_%M')
                    updated_filename = str(report_file_name).replace(f"{hours:02d}_{minutes:02d}", updated_timestamp_str)
                    report_path = Path(f'./reports/{updated_filename}')
                    logging.info(f"@@@@@@@report_path: {report_path}")
                    with open(report_path, 'rb') as f:
                        report_blob = base64.b64encode(f.read())
            try:
                workbook = openpyxl.load_workbook(report_path)
                sheet = workbook.active
                workbook_modified = openpyxl.Workbook()
                worksheet_modified = workbook_modified.active

                # Copy values and styles from the original worksheet to the new worksheet
                for row_num, row in enumerate(sheet.iter_rows(min_row=1), start=1):
                    for col_num, cell in enumerate(row, start=1):
                        new_cell = worksheet_modified.cell(row=row_num, column=col_num, value=cell.value)
                        if isinstance(cell.value, (int, float)):
                            new_cell.number_format = '0'
                workbook_modified.save(report_path)
            except Exception as e:
                logging.info(f'######{e} exception')


        else:
            with open(report_path, 'rb') as f:
                report_blob = base64.b64encode(f.read())

        try:
            logging.debug('DOWNLOADING REPORT')
            return_data = {'flag': True, 'blob': report_blob.decode('utf-8'), 'filename': f'{report_file_name}'}
        except Exception:
            message = 'Something went wrong while downloading report.'
            logging.exception(message)
            return_data = {'flag': False, 'message': message}
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except Exception:
            logging.warning("###Failed to calc end of ram and time")
            logging.exception("####ram calc went wrong####")
            memory_consumed = None
            time_consumed = None
        logging.info("## Report API download_report Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")

            
        logging.info(f"## Report API download_report Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")

        return return_data

### -------------------- from here , the code is responsible to genearate the accuarcy report ----------------


def get_total_mandatory_fields(queues_db):
    get_mad_filds="SELECT display_name,unique_name FROM `field_definition` WHERE mandatory=1"
    mad_df=queues_db.execute_(get_mad_filds)
    mandatory_fields=mad_df['unique_name'].to_list()
    total_mandatory_fields=len(mandatory_fields)
    return total_mandatory_fields,mandatory_fields

def get_edited_fields(queues_db):
    qry="SELECT `case_id`,`fields_changed` FROM `field_accuracy`"
    field_accuracy_df=queues_db.execute_(qry) 
    return field_accuracy_df

def get_ocr_fields(extraction_db,mandatory_fields,start_date,end_date):
    mad_flds_str=','.join(mandatory_fields)
    logging.info(f"### mad_flds str is {mad_flds_str}")
    qry=f"select `case_id`,{mad_flds_str} from ocr where created_date between '{start_date}' and '{end_date}'"
    ocr_df=extraction_db.execute_(qry)
    return ocr_df

### MERGE the dataframes to proceed furthur
def merge_df(df1,df2):
    merged_df = pd.merge(df1, df2, on='case_id', how='right',validate='one_to_many')
    return merged_df


MANDATORY_FIELDS = 'Mandatory Fields'
EXTRACTED_FIELDS = 'Extracted Fields'
EDITED_FIELDS = 'Edited Fields'
NOT_EXTRACTED_FIELDS = 'Not Extracted'

def generate_case_wise_accuracy(merged_df_dict,total_mandatory,mandatory_fields):
    case_accuracy={}
    for field_data in merged_df_dict:
        try:
            logging.info(f"#### field_data is {field_data['fields_changed']}")
            field_changed=field_data['fields_changed']
            case_id=field_data['case_id']
            if isinstance(field_changed, str):
                fields_changed=json.dumps(field_changed)
                field_changed=list(json.loads(json.loads(fields_changed)).keys())
                filtered_fields_changed = [field for field in field_changed if field in mandatory_fields]
                field_data_unchanged = {key: value for key, value in field_data.items() if key not in field_changed}
                condition = lambda x: x is None
                empty_vals = [value for value in field_data_unchanged.values() if condition(value)]
                not_extracted_flds=len(empty_vals)
                edited_flds=len(filtered_fields_changed)
                extract_fields_cnt=total_mandatory-not_extracted_flds
                case_accuracy[case_id]={'Mandatory Fields':total_mandatory,'Extracted Fields':extract_fields_cnt,'Edited Fields':edited_flds,'Not Extracted':not_extracted_flds}
            else:
                condition = lambda x: x is None
                # Using a list comprehension
                empty_vals = [value for value in field_data.values() if condition(value)]
                not_extracted_flds=len(empty_vals)
                extract_fields_cnt=total_mandatory-not_extracted_flds
                case_accuracy[case_id]={'Mandatory Fields':total_mandatory,'Extracted Fields':extract_fields_cnt,'Edited Fields':0,'Not Extracted':not_extracted_flds}
        except Exception as e:
            logging.info(f"### Exception occured {e}")
            continue

    return case_accuracy

#### Generate excel with the Final data
def generate_excel(data,file_name):
    df = pd.DataFrame(data).T
    summary_data={}
    mandatory_fields=df['Mandatory Fields'].to_list()
    extracted_fields=df['Extracted Fields'].to_list()
    edited_fields=df['Edited Fields'].to_list()
    not_extracted_fields=df['Not Extracted'].to_list()
    
    summary_data['Mandatory Fields']=round(sum(mandatory_fields),1)
    summary_data['Extracted Fields']=round(sum(extracted_fields),1)
    summary_data['Edited Fields']=round(sum(edited_fields),1)
    summary_data['Not Extracted']=round(sum(not_extracted_fields),1)
    accuracy = 100-(100*sum(edited_fields)/sum(mandatory_fields))
    summary_data['Accuracy']=accuracy


    logging.info(f"#### SUMMARY DATA is {summary_data}")
    df_summary = pd.DataFrame([summary_data])

    try:
        path=f'/var/www/reports_api/reports/{file_name}'
        logging.info(f"### writing to excel files path is {path}")
        with pd.ExcelWriter(path) as writer:
            df_summary.to_excel(writer, sheet_name='Summary',index=False)
            df.to_excel(writer, sheet_name='Case Wise', index_label="Case ID")

            # Access the worksheet and set column widths
            summary_sheet = writer.sheets['Summary']
            case_sheet = writer.sheets['Case Wise']

            # Set column width for 'Summary' sheet
            for i, column in enumerate(df_summary.columns):
                column_len = max(df_summary[column].astype(str).apply(len).max(), len(column))
                print(f"######## column len for summary sheet {column_len}")
                summary_sheet.set_column(i, i, column_len)

            # Set column width for 'Case Wise' sheet
            for i, column in enumerate(df.columns):
                column_len = max(df[column].astype(str).apply(len).max(), len(column))
                print(f"######## column len for case wise sheet{column_len}")
                case_sheet.set_column(i, i,20)
    except Exception as e:
        logging.error(f"### Error writing to Excel file: {e}")
    

#reconciliation_report

# =======================================================================================
# Route Name      : reconciliation_report
# Description     : This API generates the Reconciliation Report by querying necessary
#                   tables, processing the data, and returning it in both Excel format
#                   and JSON format with a structure compatible with GINGER templates.
# Inputs          : tenant_id, start_date, end_date (from request JSON)
# Output Format   : {
#                       "message": "SUCCESS",
#                       "excel_flag": 1,
#                       "flag": True,
#                       "data": [{"row_data": [...]}]
#                   }
# Excel Output    : File saved with current timestamp under /var/www/reports_api/reports/
# Usage           : Used for generating case-wise reconciliation report showing status,
#                   file names, party details, timestamps, etc.
# Note            : Converts all datetime fields to string to ensure JSON serialization.
# =======================================================================================

@app.route('/reconciliation_report', methods=['POST', 'GET'])

def reconciliation_report():
    logging.info('Calling Reconciliation Report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("Failed to start RAM/time tracking")

    try:
        data = request.json
        tenant_id = data['ui_data']['tenant_id']
        start_date = data['start_date']
        end_date = data['end_date']
        db_config['tenant_id'] = tenant_id

        trace_id = generate_random_64bit_string()
        attr = ZipkinAttrs(
            trace_id=trace_id,
            span_id=generate_random_64bit_string(),
            parent_span_id=None,
            flags=None,
            is_sampled=False,
            tenant_id=tenant_id
        )

        with zipkin_span(
            service_name='reports_api',
            zipkin_attrs=attr,
            span_name='reconciliation_report',
            transport_handler=http_transport,
            sample_rate=0.5
        ):
            hdfc_db = DB('queues', **db_config)

            report_query = f"""
                SELECT
                    NVL(TO_CHAR(pq.case_id), '') AS case_id,
                    NVL(TO_CHAR(fd.party_id), '') AS party_id,
                    NVL(TO_CHAR(pq.hub_code), '') AS hub_code,
                    NVL(TO_CHAR(pq.branch_code), '') AS branch_code,
                    NVL(TO_CHAR(pq.branch_name), '') AS branch_name,
                    NVL(TO_CHAR(ocr.party_name), '') AS party_name,
                    NVL(fd.initial_file_name, '') AS initial_file_name,
                    NVL(fd.final_file_name, '') AS final_file_name,
                    NVL(TO_CHAR(fd.merged_file_count), '') AS merged_file_count,
                    NVL(TO_CHAR(fd.file_received_datetime, 'YYYY-MM-DD HH24:MI:SS'), '') AS file_received_datetime,
                    NVL(TO_CHAR(pq.last_updated, 'YYYY-MM-DD HH24:MI:SS.FF'), '') AS last_updated,
                    NVL(TO_CHAR(pq.created_date, 'YYYY-MM-DD HH24:MI:SS'), '') AS created_date,
                    NVL(pq.status, '') AS status
                FROM hdfc_queues.file_data fd
                JOIN hdfc_extraction.ocr ocr ON fd.party_id = ocr.party_id
                JOIN hdfc_queues.process_queue pq ON fd.case_id = pq.case_id
                WHERE fd.FILE_RECEIVED_DATETIME BETWEEN 
                      TO_TIMESTAMP('{start_date}', 'YYYY-MM-DD HH24:MI:SS')
                  AND TO_TIMESTAMP('{end_date}', 'YYYY-MM-DD HH24:MI:SS')
                ORDER BY fd.FILE_RECEIVED_DATETIME DESC
            """

            df = hdfc_db.execute_(report_query)
            df.columns = df.columns.str.lower()

            df = df.loc[:, ~df.columns.duplicated()]
            df = df.drop_duplicates(subset=['initial_file_name'])

            outputs = []
            serial_number = 1
            # for _, row in df.iterrows():
            #     output = {
            #         'serial_number': serial_number,
            #         'case_id': row.get('case_id', ''),
            #         'party_id': row.get('party_id', ''),
            #         'hub_code':row.get('hub_code',''),
            #         'branch_code':row.get('branch_code',''),
            #         'branch_name':row.get('branch_name',''),
            #         'party_name': row.get('party_name', ''),
            #         'initial_file_name': row.get('initial_file_name', ''),
            #         'final_file_name': row.get('final_file_name', ''),
            #         'merged_file_count': row.get('merged_file_count', ''),
            #         'file_received_datetime': row.get('file_received_datetime', ''),
            #         'last_updated': row.get('last_updated', ''),
            #         'created_date': row.get('created_date', ''),
            #         'status': row.get('status', '')
            #     }

            #     # Convert datetime fields to string
            #     for key in ['file_received_datetime', 'last_updated', 'created_date']:
            #         if output[key]:
            #             output[key] = str(output[key])

            #     outputs.append(output)
            #     serial_number += 1

            
            required_cols = [
                'case_id', 'party_id', 'hub_code', 'branch_code', 'branch_name',
                'party_name', 'initial_file_name', 'final_file_name',
                'merged_file_count', 'file_received_datetime',
                'last_updated', 'created_date', 'status'
            ]

            # Ensure missing columns don't break code
            for col in required_cols:
                if col not in df.columns:
                    df[col] = ''

            # Keep only required columns
            df = df[required_cols]

            # Convert datetime columns to string (vectorized)
            datetime_cols = ['file_received_datetime', 'last_updated', 'created_date']
            df[datetime_cols] = df[datetime_cols].astype(str)

            # Add serial number column (fast)
            df.insert(0, 'serial_number', range(1, len(df) + 1))

            # Convert to list of dicts (fast, built-in optimized)
            outputs = df.to_dict(orient='records')
            # Save Excel file
            ist = pytz.timezone("Asia/Calcutta")
            timestamp = datetime.now(ist).strftime("%m-%d-%Y_%H_%M")
            filename = f'Reconciliation_Report--{timestamp}.xlsx'
            path = f'/var/www/reports_api/reports/{filename}'

            try:
                pd.DataFrame(outputs).to_excel(path, index=False)
            except Exception as e:
                logging.exception(f"Excel export failed: {e}")

            # Build the return JSON
            return_json_data = {
                'message': 'SUCCESS',
                'excel_flag': 1,
                'flag': True,
                'data': [{'row_data': outputs}]
            }

            # Optional: useful for template rendering
            data['report_data'] = return_json_data

    except Exception as e:
        logging.exception(f"Error generating reconciliation report: {e}")
        return jsonify({'flag': False, 'message': 'Failed to generate reconciliation report'})

    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / (1024 * 1024 * 1024)
        end_time = tt()
        logging.info(f"Memory Used: {memory_consumed:.10f} GB | Time: {round(end_time - start_time, 3)} s")
    except Exception:
        logging.warning("Memory/time tracking failed")

    return jsonify(return_json_data)








# def reconciliation_report():
#     logging.info('Calling Reconciliation Report')
#     try:
#         memory_before = measure_memory_usage()
#         start_time = tt()
#     except Exception:
#         logging.warning("Failed to start RAM/time tracking")

#     try:
#         data = request.json
#         tenant_id = data['ui_data']['tenant_id']
#         start_date = data['start_date']
#         end_date = data['end_date']
#         db_config['tenant_id'] = tenant_id

#         trace_id = generate_random_64bit_string()
#         attr = ZipkinAttrs(
#             trace_id=trace_id,
#             span_id=generate_random_64bit_string(),
#             parent_span_id=None,
#             flags=None,
#             is_sampled=False,
#             tenant_id=tenant_id
#         )

#         with zipkin_span(
#                 service_name='reports_api',
#                 zipkin_attrs=attr,
#                 span_name='reconciliation_report',
#                 transport_handler=http_transport,
#                 sample_rate=0.5
#         ):
#             hdfc_db = DB('queues', **db_config)
#             report_query = f"""
#                  SELECT
#                     NVL(TO_CHAR(pq.case_id), '') AS case_id,
#                     NVL(TO_CHAR(fd.party_id), '') AS party_id,
#                     NVL(ocr.party_name, '') AS party_name,
#                     NVL(fd.initial_file_name, '') AS initial_file_name,
#                     NVL(fd.final_file_name, '') AS final_file_name,
#                     NVL(TO_CHAR(fd.merged_file_count), '') AS merged_file_count,
#                     NVL(TO_CHAR(fd.file_received_datetime, 'YYYY-MM-DD HH24:MI:SS'), '') AS file_received_datetime,
#                     NVL(TO_CHAR(pq.last_updated, 'YYYY-MM-DD HH24:MI:SS.FF'), '') AS last_updated,
#                     NVL(TO_CHAR(pq.created_date, 'YYYY-MM-DD HH24:MI:SS'), '') AS created_date,
#                     NVL(pq.status, '') AS status
#                 FROM hdfc_queues.file_data fd
#                 JOIN hdfc_extraction.ocr ocr ON fd.party_id = ocr.party_id
#                 JOIN hdfc_queues.process_queue pq ON fd.final_file_name = pq.file_name
#                 WHERE pq.created_date BETWEEN 
#                       TO_TIMESTAMP('{start_date}', 'YYYY-MM-DD HH24:MI:SS')
#                   AND TO_TIMESTAMP('{end_date}', 'YYYY-MM-DD HH24:MI:SS')
#                 ORDER BY pq.created_date DESC
#             """
#             df = hdfc_db.execute_(report_query)
#             df.columns = df.columns.str.lower()

#             # Remove duplicate columns if any
#             df = df.loc[:, ~df.columns.duplicated()]
#             df = df.drop_duplicates(subset=['initial_file_name'])
#             #df = df.loc[:, ~df.columns.duplicated()]

#             ist = pytz.timezone("Asia/Calcutta")
#             timestamp = datetime.now(ist).strftime("%m-%d-%Y_%H_%M")
#             filename = f'Reconciliation_Report--{timestamp}.xlsx'
#             path = f'/var/www/reports_api/reports/{filename}'

#             try:
#                 df.to_excel(path, index=False)
#             except Exception as e:
#                 logging.exception(f"Excel export failed: {e}")

#             final_data = {'row_data': df.to_dict(orient='records')}

#             return_json_data = {
#                 'message': 'SUCCESS',
#                 'excel_flag': 1,
#                 'flag': True,
#                 'data': final_data,
#                 'report_data': {  # <-- added this to match accuracy report structure
#                     'message': 'Successfully excel generated'
#                 }
#             }

#     except Exception as e:
#         logging.exception(f"Error generating reconciliation report: {e}")
#         return jsonify({'flag': False, 'message': 'Failed to generate reconciliation report'})

#     try:
#         memory_after = measure_memory_usage()
#         memory_consumed = (memory_after - memory_before) / (1024 * 1024 * 1024)
#         end_time = tt()
#         logging.info(f"Memory Used: {memory_consumed:.10f} GB | Time: {round(end_time - start_time, 3)} s")
#     except Exception as e:
#         logging.warning("Memory/time tracking failed")

#     return jsonify(return_json_data)







@app.route('/generate_accuracy_report', methods=['POST', 'GET'])
def generate_accuracy_report():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
        logging.info(f"memory_before is {memory_before} and start_time is {start_time}")
    except Exception:
        logging.warning("####Failed to start ram and time calc##")
        
    data = request.json

    logging.info(f"##### Request data in GENERATE ACCURACY REPORT IS {data}")
    
    tenant_id = data.get('tenant_id', '')
    ui_data=data['ui_data']
    file_name=ui_data.get('report_output','Accurary_report')
    start_date=data['start_date']
    end_date=data['end_date']

    if start_date == end_date:
        logging.info("start and end dates are same so end date incrementing for 1 day")
        logging.info("end_date is {end_date} and type is {type(end_date)}")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        end_date= end_date.strftime("%Y-%m-%d")
        logging.info(f"end date after increment is {end_date} and type is {type(end_date)}")
    else:
        logging.info("else block executed")

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='generate_accuracy_report',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        try:
            db_config['tenant_id'] = tenant_id
            queues_db=DB('queues',**db_config)
            extraction_db=DB('extraction',**db_config)
            
            total_mandatory,mandatory_fields=get_total_mandatory_fields(queues_db)
            field_accuracy_df=get_edited_fields(queues_db)
            ocr_df=get_ocr_fields(extraction_db,mandatory_fields,start_date,end_date)
            merged_df=merge_df(field_accuracy_df,ocr_df)
            merged_df_dict=merged_df.to_dict(orient='records')

            case_wise_accuracy=generate_case_wise_accuracy(merged_df_dict,total_mandatory,mandatory_fields)
            logging.info(f"#### CASE WISE ACCURACY IS {case_wise_accuracy}")

            excel_status=generate_excel(case_wise_accuracy,file_name)
            logging.info(f"#### GENERATE EXCEL STATUS is {excel_status}")

            response_data={"flag":True,"report_data":{"message":"Successfully excel generated"}}
        except Exception as e:
            logging.info(f"### Issue while generating the accuracy report {e}")
            response_data={"flag":False,"report_data":{"message":"Excel not generated"}}

    return jsonify(response_data)


@app.route('/audit_report', methods=['POST', 'GET'])
def audit_report():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("###Failed to start ram and time calc###")
        
    data = request.json  
    tenant_id = data['ui_data']['tenant_id']
    trace_id = generate_random_64bit_string()

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ):
        data = request.json
        logging.info(f"request data: {data}")

        try:
            start_date = data['start_date']
            end_date = data['end_date']
            user=data['ui_data']['REQUESTED_BY']

            tenant_id = data['ui_data']['tenant_id']
            db_config['tenant_id'] = tenant_id

            
            query = """
                SELECT pq.case_id, ql.queue 
                FROM PROCESS_QUEUE pq 
                JOIN QUEUE_LIST ql ON ql.case_id = pq.case_id 
                WHERE CREATED_DATE >= TO_DATE(:date_param, 'YYYY-MM-DD HH24:MI:SS') 
                AND CREATED_DATE <= TO_DATE(:date_param1, 'YYYY-MM-DD HH24:MI:SS')
                AND ql.queue like '%maker%'
            """
            queue_db = DB('queues', **db_config)
            stats_db = DB('stats', **db_config)
            extraction_db = DB("extraction", **db_config)

            params = {'date_param': start_date, 'date_param1': end_date}
            df = queue_db.execute_(query, params=params)

            case_ids = df['case_id'].tolist()
            logging.info(f"case_ids{case_ids}")

            outputs = []
        
            for serial_number, case_id in enumerate(case_ids, start=1):
                audit_query = f"SELECT case_id, updated_date, ingested_queue, api_service FROM audit_ WHERE case_id LIKE '%{case_id}%'  AND api_service IN ('create_case_id', 'update_queue')"
                audit_case_data = stats_db.execute_(audit_query)
                
                logging.info(f"case_id{case_id}")
                
                grouped = audit_case_data.groupby('ingested_queue')
                
            
                query_file_recieved = f"SELECT CREATED_DATE from PROCESS_QUEUE where case_id like '%{case_id}%'"
                df_file_recieved = queue_db.execute_(query_file_recieved)
                logging.info(f"df_file_recieved{df_file_recieved}")
                query_party = f"SELECT PARTY_ID , PARTY_NAME , MOVED_BY from OCR where case_id like '%{case_id}%'"
                df_party = extraction_db.execute_(query_party)
                logging.info(f"df_party{df_party}")
                try:
                    party_id=df_party['party_id'][0]
                    party_name = df_party['party_name'][0]
                    moved_by = df_party['MOVED_BY'][0]
                except Exception:
                    party_id=None
                    party_name = None
                    moved_by = None


                if not df_file_recieved.empty:
                    desired_updated_date = df_file_recieved['created_date'].iloc[0]
                else:
                    desired_updated_date=Timestamp('00:00:00')

                
                # Calculate relevant timestamps and time differences
                try:
                    file_received_time = desired_updated_date
                except Exception:
                    pass
                try:
                    maker_ingestion_time = grouped.get_group('maker_queue')['updated_date'].min()
                except Exception:
                    pass
            
                try:
                    completed_ingested_time = grouped.get_group('accepted_queue')['updated_date'].min()
                except Exception:
                    
                    completed_ingested_time = '0000-00-00 00:00:00'
                    
            

                try:
                    rejected_ingested_time = grouped.get_group('rejected_queue')['updated_date'].min()
                except Exception:
                    
                    rejected_ingested_time = '0000-00-00 00:00:00'
                    
            
                

                if rejected_ingested_time == '0000-00-00 00:00:00':
                    try:
                        
                        Total_handling_time= completed_ingested_time - maker_ingestion_time
                    except Exception:
                        Total_handling_time = '0000-00-00 00:00:00'
                else:
                    try:
                        
                        Total_handling_time=rejected_ingested_time - maker_ingestion_time
                    except Exception:
                        Total_handling_time = '0000-00-00 00:00:00'


                
                TOTAL_HANDLING_TIME = 'Total Handling time'
        
                output = {
                        'serial_number': serial_number,
                    
                        'case_id': case_id,
                        'Party_id':party_id,
                        'Party_name':party_name,
                        'Business Rules Validation' : 'NA',
                        'CLIMS Create API Request time':'NA',
                        'CLIMS Create API Response Time':'NA',
                        
                        'Case creation time stamp': file_received_time,
            
                        'Maker queue in Time Stamp': maker_ingestion_time,
                        'Completed queue in time Stamp':completed_ingested_time,
                        'Rejected queue in time Stamp':rejected_ingested_time,
                        'Total Handling time':Total_handling_time,
                        'Maker Name':moved_by,
                        'User ID':user

                        
                        
                        
                        
                    }
                outputs.append(output)
                logging.info(f"output{output}")
            
            logging.info(f"outputs##{outputs}")

            SUCCESS_MESSAGE_AR = 'Successfully generated the report'

            for output in outputs:
                output['Case creation time stamp'] = str(output['Case creation time stamp'])
                output['Maker queue in Time Stamp'] = str(output['Maker queue in Time Stamp'])
                output['Completed queue in time Stamp'] = str(output['Completed queue in time Stamp'])
                output['Rejected queue in time Stamp'] = str(output['Rejected queue in time Stamp'])
                output['Total Handling time'] = str(output['Total Handling time'])
            return_json_data = {}
            logging.info(f'{return_json_data}###return_json_data#########return_json_data')
            return_json_data['message']='SUCCESS_MESSAGE_AR'
            return_json_data['excel_flag']= 1
            return_json_data['flag'] = True
            return_json_data['data'] = [{'row_data':outputs}]
            data['report_data'] = return_json_data

            logging.info(f'{return_json_data}###############return_json_data')


            FAILED_MESSAGE_AR = 'Failed!!'
            
        except Exception as e:
            logging.info(f"error at audit_Report {e}")
            logging.debug(f"{e} ####issue")
            return_json_data = {}
            return_json_data['flag'] = False
            return_json_data['message'] = 'FAILED_MESSAGE_AR'
            return jsonify(return_json_data)
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except Exception:
        logging.warning("####Failed to calc end of ram and time")
        logging.exception("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        

    # insert audit
    NEW_FILE_RECEIVED_AR = "New file received"
    try:
        audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
                        "api_service": "folder_monitor", "service_container": "reportsapi",
                        "changed_data": NEW_FILE_RECEIVED_AR,"tables_involved": "","memory_usage_gb": str(memory_consumed), 
                        "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                        "response_data": json.dumps(return_json_data), "trace_id": trace_id,
                        "session_id": "","status":json.dumps(return_json_data['flag'])}
  
        insert_into_audit(audit_data)
    except Exception:
        logging.info("#issue in the query formation#")
    return jsonify(return_json_data)



    
def extract_numeric_value(value):
    if isinstance(value, str):
        cleaned_value = re.sub(r'[^\d.]', '', value)  # Remove non-numeric characters except dot
        return pd.to_numeric(cleaned_value, errors='coerce') if cleaned_value else None
    return None

def replace_empty_with_none(value):
    return None if value == "" else value

def create_dataframe_from_json(json_data_list):
    columns_data = {}

    try:
        for idx, json_data in enumerate(json_data_list):
            if json_data is not None and isinstance(json_data, dict):  # Check if json_data is not None and is a dictionary
                for key, value in json_data.items():
                    if value is not None and value != 'null':
                        try:
                            data_dict = json.loads(value)
                            for nested_key, nested_value in data_dict.items():
                                column_name = f"{nested_key}"
                                if column_name not in columns_data:
                                    columns_data[column_name] = []
                                numeric_value = extract_numeric_value(replace_empty_with_none(nested_value))
                                columns_data[column_name].append(numeric_value)
                        except json.JSONDecodeError as json_error:
                            logging.info(f"Error decoding JSON at index {idx}: {json_error}")
                            # Handle the error if needed
                    else:
                        # Handle the case where value is None
                        for nested_key in columns_data.keys():
                            columns_data[nested_key].append(None)
            else:
                # Handle the case where json_data is None or not a dictionary
                print(f"Skipping invalid data at index {idx}: {json_data}")
                for nested_key in columns_data.keys():
                    columns_data[nested_key].append(None)

        # Fill missing keys with None
        if len(columns_data) !=0:
            max_len = max(len(v) for v in columns_data.values())
            for key, values in columns_data.items():
                values.extend([None] * (max_len - len(values)))

        df = pd.DataFrame(columns_data)
        return df
    except Exception as e:
        logging.info(f"ERROR : {e}")
        return None


def convert_to_custom_format(date_str):
    formats = [
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%b %Y",
        "%d %b %Y",
        "%B %Y",
        "%m-%Y",
        "%m/%Y",
        "%d %B, %Y",
        "%B-%Y",
    ]

    for date_format in formats:
        try:
            dt = datetime.strptime(date_str, date_format)
            return dt.strftime("%d-%b-%y")
        except ValueError as e:
            logging.info(f"{e}###wrong format")

    return None

def clean_comments(comment_json):
    try:
        comment_list = json.loads(comment_json)
        # Extract 'comment' field, replace newlines with commas
        return ', '.join(comment_list[0]['comment'].splitlines())
    except Exception as e:
        return ''  # fallback if parsing fails


@app.route('/consolidated_report', methods=['POST', 'GET'])
def consolidated_report():
    logging.info('Calling Consolidated report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("#Failed to start ram and time calc#")
        
    data = request.json
    tenant_id = data['ui_data']['tenant_id']
        
    
    trace_id = generate_random_64bit_string()

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ):
        data=request.json
        tenant_id = data['ui_data']['tenant_id']
        db_config['tenant_id'] = tenant_id
        logging.info(f"Request Data is: {data}")

        reports_name=data['ui_data']['report_name']
        extraction_db = DB('extraction', **db_config)

        queues_db = DB("queues", **db_config)

        start_date = data['start_date']
        end_date = data['end_date']
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)

        reports_name=data['ui_data']['report_name']
        extraction_db = DB('extraction', **db_config)
        
        queues_db = DB("queues", **db_config)
        

        start_date = data['start_date']
        end_date = data['end_date']
        
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)

        timestamp1 = timestamp.strftime('%d-%m-%Y %H:%M:%S')

        timestamp = str(timestamp)[:-13]
        logging.info("reports_name{reports_name}")
        parsed_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
        formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")

        try:
            if 'completed' in reports_name.lower():
                file='completed'
                query_case_ids=f""" SELECT pq.case_id, o.party_id, pq.hub_code, pq.branch_code, pq.branch_name, o.customer_name, pq.region, o.rm_state, o.relation_mgrname, pq.created_date, pq.last_updated_by, o.last_updated, pq.status, o.remarks FROM hdfc_queues.process_queue pq INNER JOIN hdfc_queues.queue_list ql ON pq.case_id = ql.case_id INNER JOIN hdfc_extraction.ocr o
            ON pq.case_id = o.case_id WHERE (ql.queue = 'accepted_queue') AND pq.last_updated BETWEEN TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            
            
            if 'rejected' in reports_name.lower():
                file='rejected'
                query_case_ids=f""" SELECT pq.case_id, o.party_id,pq.hub_code, pq.branch_code, pq.branch_name, o.customer_name, pq.region, o.rm_state, o.relation_mgrname, pq.created_date, pq.last_updated_by, o.last_updated, pq.status, o.rejected_comments,o.waiting_queue_rejected_comments FROM hdfc_queues.process_queue pq INNER JOIN hdfc_queues.queue_list ql ON pq.case_id = ql.case_id INNER JOIN hdfc_extraction.ocr o
            ON pq.case_id = o.case_id WHERE (ql.queue = 'rejected_queue') AND pq.last_updated BETWEEN TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            
            
            if 'onhold' in reports_name.lower():
                logging.info("entered hold report block")
                file='onhold'
                query_case_ids=f""" SELECT pq.case_id, o.party_id,pq.hub_code, pq.branch_code, pq.branch_name, o.customer_name, pq.region, o.rm_state, o.relation_mgrname,  pq.created_date, pq.last_updated_by, o.last_updated, pq.status, o.comments FROM hdfc_queues.process_queue pq INNER JOIN hdfc_queues.queue_list ql ON pq.case_id = ql.case_id INNER JOIN hdfc_extraction.ocr o
            ON pq.case_id = o.case_id WHERE pq.status = 'Hold' AND pq.last_updated BETWEEN TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            
        
            filename = f'{file}_report--{formatted_date}.xlsx'
            output_file=f'/var/www/reports_api/reports/{filename}'
            files_df=queues_db.execute_(query_case_ids)
            files_df = files_df.loc[:, ~files_df.columns.str.lower().duplicated()]
            if 'onhold' in reports_name.lower():
                files_df['comments'] = files_df['comments'].apply(clean_comments)
            if 'rejected' in reports_name.lower():
                files_df['waiting_queue_rejected_comments'] = files_df['waiting_queue_rejected_comments'].apply(clean_comments)
            files_df.to_excel(output_file,index=False)
            
            
            columns=files_df.columns
            files_df_dict=files_df.to_dict(orient='records')

            return_json_data = {}
            return_json_data['message']='Successfully generated the report'
            return_json_data['flag'] = True
            return_json_data['excel_flag']= 1
            return_json_data['data'] = {'row_data':files_df_dict}
            data['report_data'] = return_json_data
            return jsonify(return_json_data)
        except Exception as e:
            logging.exception(f"Expected Occured so returing empty file with columns")
            columns = ["Case ID", "Party ID", "HUB Code" , "Branch Code" , "Branch Name","Customer Name", "Region","RM State","Relation MGR Name","Created Date & Time","Last Updated By","Last Updated","Status","Comments"]
            df = pd.DataFrame(columns=columns)
            df.to_excel(output_file,index=False)
            return_json_data = {}
            return_json_data['flag'] = False
            return_json_data['message'] = 'Failed!!'
            return jsonify(return_json_data)

        
              


    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except Exception:
        logging.warning("######Failed to calc end of ram and time")
        logging.exception("##ram calc went wrong")
        memory_consumed = None
        time_consumed = None
    try:
        audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
                    "api_service": "consolidated_report", "service_container": "reportsapi",
                    "changed_data":"report_downloaded","tables_involved": "","memory_usage_gb": str(memory_consumed), 
                    "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                    "response_data": json.dumps(return_json_data), "trace_id": trace_id,
                    "session_id": "","status":json.dumps(return_json_data['flag'])}
    
        insert_into_audit(audit_data)
    except Exception:
        logging.exception("issue in the query formation")
    return jsonify(return_json_data)








@app.route('/role_master_report', methods=['POST', 'GET'])
def role_master_report():
    logging.info('Calling Role Master Report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("#Failed to start ram and time calc#")
        
    data = request.json
    tenant_id = data['ui_data']['tenant_id']
        
    
    trace_id = generate_random_64bit_string()

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ):
        tenant_id = data['ui_data']['tenant_id']
        db_config['tenant_id'] = tenant_id
        logging.info(f"Request Data is: {data}")

        reports_name=data['ui_data']['report_name']
        group_access_db = DB('group_access', **db_config)

        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        timestamp = str(timestamp)[:-13]

        try:
            role_rights=f"select role_name,role_description,type_of_access,profile_assigned_to,display_role_rights,new_rights_assigned_status from role_rights"
            role_rights_df=group_access_db.execute_(role_rights).to_dict(orient="records")

            columns = ['Role Name','Role Description','Rights Assigned to Role','Rights Description','Rights Assigned Y/N','Status of the Role Name','Profile Assigned To','Type of Access','Creation Date','Disable Date']
            res = pd.DataFrame(columns=columns)
            rights_description= {
                "Add User": "Adding the new user into the application",
                "Modify User": "Modify the existing user details in the application",
                "Add Business Rule": "Adding business rules to the application",
                "Modify Business Rule": "Modify the existing business rules in the application",
                "Add Roles": "Create new roles",
                "Modify Roles":"Modify the existing roles",
                "View All Queues": "View all queues in the application",
                "Modify All Queues": "Editing the details of the master data",
                "Master Data Edit": "Editing the details of the master data",
                "Bulk Transaction": "Managing a large volume of transactions simultaneously",
                "Approve UAM Maker Activity": "Approve added new user, modification, deletion, deactivation and enable existing user",
                "Reject UAM Maker Activity": "Reject added new user, modification, deletion, deactivation and enable existing user",
                "Approve edits to Master": "Approve edits to Master",
                "Reject edit to Master": "Reject edit to Master",
                "Approve change in Business Rule": "Approve change in Business Rule",
                "Reject change in Business Rule": "Reject change in Business Rule",
                "Operation Reports": "Access to operations reports",
                "UAM Reports":"Access to uamreports"
                }
            logging.info(f'rights_description:{rights_description}')
            def get_role_details(role_rights_df, role_name):
                result = [entry for entry in role_rights_df if entry["role_name"] == role_name]
                return result
            role_names=f"select group_name,status,created_date,disabled_date from group_definition"
            group_definition_data=group_access_db.execute_(role_names)
            role_names_list=group_definition_data['group_name'].tolist()
            group_definition_dict = {entry["group_name"]: entry for entry in group_definition_data.to_dict(orient="records")}
            
            for role_name in role_names_list:
                one_role_rights=get_role_details(role_rights_df,role_name)
                logging.info(f"one_role_rights:{one_role_rights}")
                group_info = group_definition_dict.get(role_name, {})
                status_= group_info.get("status", "Unknown")
                created_date = group_info.get("created_date", None)
                disabled_date = group_info.get("disabled_date", None)
                for entry in one_role_rights:
                    logging.info(f"###entry :{entry}")
                    rights_assigned = entry.get("display_role_rights", "").split(",")  # List of rights
                    rights_statuses = entry.get("new_rights_assigned_status", "").split(",")  # List of statuses
                    role_description= entry.get("role_description", None)
                    type_of_access= entry.get("type_of_access", None)
                    profile_assigned_to= entry.get("profile_assigned_to", None)

                    for right, status in zip(rights_assigned, rights_statuses):  # Pair each right with its status
                        res = res.append({
                            'Role Name': role_name,
                            'Role Description':role_description,
                            'Rights Assigned to Role': right.strip(),
                            'Rights Description': rights_description.get(right.strip(), ""),
                            'Rights Assigned Y/N': 'Y' if status.strip().lower() == 'yes' else 'N',
                            'Status of the Role Name': 'Active' if status_ == "enabled" else 'Inactive',
                            'Profile Assigned To':profile_assigned_to,
                            'Type of Access':type_of_access,
                            'Creation Date': created_date,
                            'Disable Date':"" if status_ == "enabled" else disabled_date
                        }, ignore_index=True)


            parsed_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            filename = f'Role_Master_Report--{formatted_date}.xlsx'
            path = f'/var/www/reports_api/reports/{filename}'
            try:

                res.to_excel(path, index=False)
            except Exception as e:
                
                logging.info(f"######in exception for excel conversion error is {e}")     
            given_data=res.to_dict()
            final_data = {'row_data': []}
    

            # Iterating through each entry in the given data dictionary
            logging.info(f"#given data{given_data}")
            if given_data:
                for i in range(len(given_data['Role Name'])):
                    row_entry = {
                    'Role Name':given_data['Role Name'][i],
                    'Role Description':given_data['Role Description'][i],
                    'Rights Assigned to Role':given_data['Rights Assigned to Role'][i],
                    'Rights Description':given_data['Rights Description'][i],
                    'Rights Assigned Y/N':given_data['Rights Assigned Y/N'][i],
                    'Status of the Role Name':given_data['Status of the Role Name'][i],
                    'Profile Assigned To':given_data['Profile Assigned To'][i],
                    'Type of Access':given_data['Type of Access'][i],
                    'Creation Date': (
                        given_data['Creation Date'][i].strftime("%d-%b-%y %H:%M:%S")
                        if isinstance(given_data['Creation Date'][i], (datetime, pd.Timestamp)) and not pd.isna(given_data['Creation Date'][i])
                        else ""
                    ),
                    'Disable Date': (
                        given_data['Disable Date'][i].strftime("%d-%b-%y %H:%M:%S")
                        if isinstance(given_data['Disable Date'][i], (datetime, pd.Timestamp)) and not pd.isna(given_data['Disable Date'][i])
                        else ""
                    )        
                    }
                    final_data['row_data'].append(row_entry)
            logging.info(f"#final data{final_data}")


            try: 
                SUCCESS_MESSAGE_PRA='Successfully generated the report'

                return_json_data = {}
                return_json_data['message']='SUCCESS_MESSAGE_PRA'
                return_json_data['excel_flag']= 1
                return_json_data['flag'] = True
                return_json_data['data'] = final_data
                data['report_data'] = return_json_data

                logging.info(f'{return_json_data}###############return_json_data')
                try:
                    json.dumps(return_json_data)  # Test serialization
                except TypeError as e:
                    print(f"Serialization error: {e}")            
                
                FAILED_MESSAGE_PRA='Failed!!'

            except Exception as e:
                logging.info(f"error at process_Report {e}")
                logging.debug(f"{e} ####issue")
                return_json_data = {}
                return_json_data['flag'] = False
                return_json_data['message'] = ' FAILED_MESSAGE_PRA'
                
        except Exception as e:
            logging.exception(f'Something went wrong exporting data : {e}')
            return {'flag': False, 'message': 'Unable to export data.'}
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except Exception:
        logging.warning("#####Failed to calc end of ram and time")
        logging.exception("#ram calc went wrong#")
        memory_consumed = None
        time_consumed = None
        

    # # insert audit
    # NEW_FILE_RECEIVED_CR = "New file received"

    # try:
    #     audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
    #                     "api_service": "consolidated_report", "service_container": "reportsapi",
    #                     "changed_data": NEW_FILE_RECEIVED_CR,"tables_involved": "","memory_usage_gb": str(memory_consumed), 
    #                     "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
    #                     "response_data": json.dumps(return_json_data), "trace_id": trace_id,
    #                     "session_id": "","status":json.dumps(return_json_data['flag'])}
        
    #     insert_into_audit(audit_data)
    # except Exception:
    #     logging.info("##issue in the query formation##")
    return jsonify(return_json_data)


@app.route('/user_id_population_report', methods=['POST', 'GET'])
def user_id_population_report():
    logging.info('Calling User Id Population Report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("#Failed to start ram and time calc#")
       
    data = request.json
       
   
    trace_id = generate_random_64bit_string()
 
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id='hdfc'
    )
 
    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ):
        #tenant_id = data['tenant_id']
        db_config['tenant_id'] = 'hdfc'
        logging.info(f"Request Data is: {data}")

        report_name = data.get('report_name', None)
        generated_datetime = data.get('generated_datetime')
        group_access_db = DB('group_access', **db_config)
 
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        timestamp = str(timestamp)[:-13]
 
        try:
            user_id_population_report_query=f"""            
                        SELECT
                        NVL(AD.USERNAME, ' ') AS USERNAME,
                        NVL(AD.EMPLOYEE_NAME, ' ') AS EMPLOYEE_NAME,
                        NVL(AD.EMPLOYEE_CODE, ' ') AS EMPLOYEE_CODE,
                        NVL(AD.BRANCH_CODE, ' ') AS BRANCH_CODE,
                        NVL(AD.BRANCH_NAME, ' ') AS BRANCH_NAME,
                        NVL(AD.DEPARTMENT_CODE, ' ') AS DEPARTMENT_CODE,
                        NVL(AD.DEPARTMENT_NAME, ' ') AS DEPARTMENT_NAME,
                        NVL(AD.ROLE, ' ') AS ROLE,
                        NVL(
                            CASE
                                WHEN LOWER(AD.STATUS) = 'disable' THEN 'Disabled'
                                WHEN LOWER(AD.STATUS) = 'lock' THEN 'Locked'
                                ELSE INITCAP(AD.STATUS)
                            END, ' ') AS STATUS,
                        NVL(TO_CHAR(AD.CREATED_DATE, 'DD-MM-YYYY HH24:MI:SS'), ' ') AS CREATED_DATE,
                        NVL(TO_CHAR(MAX(LL.LOGIN), 'DD-MM-YYYY HH24:MI:SS'), ' ') AS LAST_LOGIN,
                        CASE  
                            WHEN AD.MAKER_ID IS NULL AND AD.MAKER_NAME IS NULL THEN ' '
                            ELSE NVL(AD.MAKER_ID, ' ') || ',' || NVL(AD.MAKER_NAME, ' ')
                        END AS MAKER_ID_NAME,
                        NVL(TO_CHAR(AD.MAKER_DATE, 'DD-MM-YYYY HH24:MI:SS'),' ') AS MAKER_DATE,
                        CASE  
                            WHEN AD.CHECKER_ID IS NULL AND AD.CHECKER_NAME IS NULL THEN ' '
                            ELSE NVL(AD.CHECKER_ID, ' ') || ',' || NVL(AD.CHECKER_NAME, ' ')
                        END AS CHECKER_ID_NAME,
                        NVL(TO_CHAR(AD.CHECKER_DATE, 'DD-MM-YYYY HH24:MI:SS'), ' ') AS CHECKER_DATE,
                        NVL(TO_CHAR(AD.USER_DISABLED_DATE, 'DD-MM-YYYY HH24:MI:SS'), ' ') AS USER_DISABLED_DATE,
                        NVL(TO_CHAR(AD.DELETED_DATE, 'DD-MM-YYYY HH24:MI:SS'), ' ') AS DELETED_DATE
                    FROM  
                        hdfc_group_access.ACTIVE_DIRECTORY AD
                    LEFT JOIN  
                        hdfc_group_access.LOGIN_LOGOUT LL  
                        ON AD.USERNAME = LL.USER_NAME
                    WHERE  
                        AD.STATUS NOT IN ('waiting','rejected','closed')
                    GROUP BY
                        AD.USERNAME,
                        AD.EMPLOYEE_NAME,
                        AD.EMPLOYEE_CODE,
                        AD.BRANCH_CODE,
                        AD.BRANCH_NAME,
                        AD.DEPARTMENT_CODE,
                        AD.DEPARTMENT_NAME,
                        AD.ROLE,
                        AD.STATUS,
                        AD.CREATED_DATE,
                        AD.MAKER_ID,
                        AD.MAKER_NAME,
                        AD.MAKER_DATE,
                        AD.CHECKER_ID,
                        AD.CHECKER_NAME,
                        AD.CHECKER_DATE,
                        AD.USER_DISABLED_DATE,
                        AD.DELETED_DATE
                    ORDER BY
                        AD.CHECKER_DATE DESC, AD.MAKER_DATE DESC"""
            user_id_population_report_df=group_access_db.execute_(user_id_population_report_query).to_dict(orient="records")
            logging.info(f"user_id_population_report_df : {user_id_population_report_df}")
            columns= ["User ID", "User Name", "Employee Code", "Branch Code", "Branch Name", "Department Code", "Department Name", "Profile 1", "Status 1", "Creation Date & Time", "Last Password Changed Date & Time", "Last Log in Date & Time", "Last Modified Maker ID and Maker Name", "Maker Date & Time", "Last Modified Checker ID and Checker Name", "Checker Date & Time", "Disabled Date & Time", "Profile End Date & Time"]
            df = pd.DataFrame(columns=columns)

            def parse_date(date_str):
                if pd.isna(date_str):
                    return pd.NaT
                try:
                    # val = datetime.strptime(date_str, "%d-%b-%Y %I:%M:%S %p") # worked for some cases
                    if 'AM' in date_str or 'PM' in date_str:
                        val = datetime.strptime(date_str, "%d-%b-%Y %I:%M:%S %p")
                    else:
                        # If no AM/PM, use the 24-hour format
                        val = datetime.strptime(date_str, "%d-%m-%Y %H:%M:%S")
                    return val
                except Exception as e:
                    print(f"Except occured {e}")
                    pass
                    return pd.NaT

            # Append the records to the DataFrame
            for record in user_id_population_report_df:
                df = df.append({
                    "User ID": record.get("USERNAME", ""),
                    "User Name": record.get("EMPLOYEE_NAME", ""),
                    "Employee Code": record.get("EMPLOYEE_CODE", ""),
                    "Branch Code": record.get("BRANCH_CODE", ""),
                    "Branch Name": record.get("BRANCH_NAME", ""),
                    "Department Code": record.get("DEPARTMENT_CODE", ""),
                    "Department Name": record.get("DEPARTMENT_NAME", ""),
                    "Profile 1": record.get("ROLE", ""),
                    "Status 1": record.get("STATUS", ""),
                    "Creation Date & Time": record.get("CREATED_DATE", ""),
                    "Last Password Changed Date & Time": "",  # No password change field in the query
                    "Last Log in Date & Time": record.get("LAST_LOGIN", ""),
                    "Last Modified Maker ID and Maker Name": record.get("MAKER_ID_NAME", ""),
                    "Maker Date & Time": record.get("MAKER_DATE", ""),
                    "Last Modified Checker ID and Checker Name": record.get("CHECKER_ID_NAME", ""),
                    "Checker Date & Time": record.get("CHECKER_DATE", ""),
                    "Disabled Date & Time": record.get("USER_DISABLED_DATE", ""),
                    "Profile End Date & Time": record.get("DELETED_DATE", "")
                }, ignore_index=True)

            parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'User_Id_Population_report--{formatted_date}.xlsx'
            #df = pd.read_excel(/var/www/reports_api/reports/)

            df['Maker Date & Time'] = df['Maker Date & Time'].astype(str).str.strip()
            df['Maker Date & Time'] = df['Maker Date & Time'].apply(parse_date)

            df['Creation Date & Time'] = df['Creation Date & Time'].astype(str).str.strip()
            df['Creation Date & Time'] = df['Creation Date & Time'].apply(parse_date)

            df['Last Password Changed Date & Time'] = df['Last Password Changed Date & Time'].astype(str).str.strip()
            df['Last Password Changed Date & Time'] = df['Last Password Changed Date & Time'].apply(parse_date)

            df['Last Log in Date & Time'] = df['Last Log in Date & Time'].astype(str).str.strip()
            df['Last Log in Date & Time'] = df['Last Log in Date & Time'].apply(parse_date)
            
            df['Checker Date & Time'] = df['Checker Date & Time'].astype(str).str.strip()
            df['Checker Date & Time'] = df['Checker Date & Time'].apply(parse_date)

            df['Disabled Date & Time'] = df['Disabled Date & Time'].astype(str).str.strip()
            df['Disabled Date & Time'] = df['Disabled Date & Time'].apply(parse_date)

            df['Profile End Date & Time'] = df['Profile End Date & Time'].astype(str).str.strip()
            df['Profile End Date & Time'] = df['Profile End Date & Time'].apply(parse_date)

            df.to_excel(f'/var/www/reports_api/reports/{report_file_name}', index=False)


            

            output_file=f'/var/www/reports_api/reports/{report_file_name}'
            
            df = pd.read_excel(output_file)

            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="Sheet1", index=False)
                
                # Load workbook and sheet
                workbook = writer.book
                sheet = workbook["Sheet1"]
                # Define a date format style
                date_style = NamedStyle(name="datetime", number_format="DD-MM-YYYY HH:MM:SS")
                # date_columns = ["Creation Date & Time", "Last Log in Date & Time", "Maker Date & Time","Checker Date & Time","Disabled Date & Time","Profile End Date & Time"]

                # # Apply style to 'Creation Date & Time' column (B column in this case)
                for cell in sheet["R"][1:]:  # Skip header row
                    cell.style = date_style
                for cell in sheet["N"][1:]:  # Skip header row
                    cell.style = date_style
                for cell in sheet["J"][1:]:  # Skip header row
                    cell.style = date_style
                for cell in sheet["L"][1:]:  # Skip header row
                    cell.style = date_style
                for cell in sheet["P"][1:]:  # Skip header row
                    cell.style = date_style
                for cell in sheet["Q"][1:]:  # Skip header row
                    cell.style = date_style
 
                workbook.save(output_file)

            wb = load_workbook(output_file)
            ws = wb.active
                
            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get column letter (A, B, C, etc.)

                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = max_length + 2  # Add padding
                ws.column_dimensions[col_letter].width = adjusted_width

            # Save the adjusted file
            wb.save(output_file)
            try:
                with open(output_file, 'rb') as f:
                    report_blob = base64.b64encode(f.read())
            except FileNotFoundError:
                logging.error(f"File not found: {output_file}")
                print({'flag': False, 'message': f"File not found: {output_file}"})
            except Exception as e:
                logging.error(f"Failed to read file {output_file}: {e}")
                print({'flag': False, 'message': "Failed to read report file."})
            
        except Exception as e:
            logging.exception(f'Something went wrong exporting data : {e}')
            return {'flag': False, 'message': 'Unable to export data.'}
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except Exception:
        logging.warning("#####Failed to calc end of ram and time")
        logging.exception("#ram calc went wrong#")
        memory_consumed = None
        time_consumed = None
       
 
    # # insert audit
    # NEW_FILE_RECEIVED_CR = "New file received"
 
    # try:
    #     audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
    #                     "api_service": "consolidated_report", "service_container": "reportsapi",
    #                     "changed_data": NEW_FILE_RECEIVED_CR,"tables_involved": "","memory_usage_gb": str(memory_consumed),
    #                     "time_consumed_secs": time_consumed, "request_payload": json.dumps(data),
    #                     "response_data": json.dumps(return_json_data), "trace_id": trace_id,
    #                     "session_id": "","status":json.dumps(return_json_data['flag'])}
       
    #     insert_into_audit(audit_data)
    # except Exception:
    #     logging.info("##issue in the query formation##")
    return {'flag': True, 'blob': report_blob.decode('utf-8'), 'filename': f'{report_file_name}'}
 

@app.route('/process_report_agri', methods=['POST', 'GET'])
def process_report_agri():
    """
        This route is responsible to generate report based on the cases processed 
        This route used to take late of time if case volume is high so changed the logic
        to only read the required fields form the Application
    """
    logging.info('Calling agri process report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("Failed to start ram and time calc")
        
    data = request.json
    trace_id = generate_random_64bit_string()
    tenant_id = data['ui_data']['tenant_id']

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ):
        data = request.json
        tenant_id = data['ui_data']['tenant_id']
        db_config['tenant_id'] = tenant_id
        logging.info(f"Request Data is: {data}")

        extraction_db = DB('extraction', **db_config)
        queues_db = DB('queues', **db_config)
        
        start_date = data['start_date']
        end_date = data['end_date']
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        
        

        timestamp = str(timestamp)[:-13]

        parsed_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
        formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
        filename = f'Process_Report--{formatted_date}.xlsx'
        output_file = f'/var/www/reports_api/reports/{filename}'
        try:
            # query_case_ids  =f"""SELECT PQ.CASE_ID,O.PARTY_ID,O.CUSTOMER_NAME,PQ.REGION,O.RM_STATE,O.RELATION_MGRNAME,O.STOCK_DOC_MONTH,O.STOCK_DOC_YEAR, PQ.CREATED_DATE,PQ.LAST_UPDATED_BY, O.LAST_UPDATED, PQ.STATUS FROM HDFC_QUEUES.PROCESS_QUEUE PQ INNER JOIN HDFC_QUEUES.QUEUE_LIST QL ON PQ.CASE_ID = QL.CASE_ID INNER JOIN HDFC_EXTRACTION.OCR O 
            # ON PQ.CASE_ID = O.CASE_ID WHERE (QL.QUEUE = 'accepted_queue' OR QL.QUEUE LIKE '%maker%')  AND PQ.CREATED_DATE >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND PQ.CREATED_DATE <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            query_case_ids=f""" SELECT pq.case_id, o.party_id,pq.hub_code, pq.branch_code, pq.branch_name, o.customer_name, pq.region, o.rm_state, o.relation_mgrname, o.stock_doc_month, o.stock_doc_year, pq.created_date, pq.last_updated_by, o.last_updated, pq.status
            FROM hdfc_queues.process_queue pq INNER JOIN hdfc_queues.queue_list ql ON pq.case_id = ql.case_id INNER JOIN hdfc_extraction.ocr o
            ON pq.case_id = o.case_id WHERE (ql.queue = 'accepted_queue' OR ql.queue LIKE '%maker%' OR ql.queue = 'rejected_queue' OR ql.queue = 'waiting_queue') AND pq.created_date BETWEEN TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            files_df=queues_db.execute_(query_case_ids)
            # files_df.rename(columns={
            #     'CASE_ID': 'ACE Case ID',
            #     'PARTY_ID': 'Party ID',
            #     'PARTY_NAME': 'Party name'}, inplace=True)
            
            # Drop the specified columns
            # columns_to_drop = ["CASE_ID", "PARTY_ID", "CUSTOMER_NAME", "REGION", "RM_STATE", "RELATION_MGRNAME", "STOCK_DOC_MONTH", "STOCK_DOC_YEAR", "CREATED_DATE", "LAST_UPDATED_BY", "LAST_UPDATED", "STATUS"]
            # columns_to_drop_existing = [col for col in columns_to_drop if col in files_df.columns]
            files_df = files_df.loc[:, ~files_df.columns.str.lower().duplicated()]
            files_df.to_excel(output_file,index=False)
            
            
            columns=files_df.columns
            files_df_dict=files_df.to_dict(orient='records')

            return_json_data = {}
            return_json_data['message']='Successfully generated the report'
            return_json_data['flag'] = True
            return_json_data['excel_flag']= 1
            return_json_data['data'] = {'row_data':files_df_dict}
            data['report_data'] = return_json_data
            return jsonify(return_json_data)
        except Exception as e:
            logging.exception(f"Expected Occured so returing empty file with columns")
            columns = ["ACE Case ID", "Party ID","HUB Code","Branch Code","Branch Name" ,"Party name", "Region","RM State","Relation MGR Name","Stock Doc Month","Stock Doc Year","Created Date & Time","Modified By","Last Updated","Status"]
            df = pd.DataFrame(columns=columns)
            df.to_excel(output_file,index=False)
            return_json_data = {}
            return_json_data['flag'] = False
            return_json_data['message'] = 'Failed!!'
            return jsonify(return_json_data)

        
              


    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except Exception:
        logging.warning("######Failed to calc end of ram and time")
        logging.exception("##ram calc went wrong")
        memory_consumed = None
        time_consumed = None
    try:
        audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
                    "api_service": "process_report_agri", "service_container": "reportsapi",
                    "changed_data":"report_downloaded","tables_involved": "","memory_usage_gb": str(memory_consumed), 
                    "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                    "response_data": json.dumps(return_json_data), "trace_id": trace_id,
                    "session_id": "","status":json.dumps(return_json_data['flag'])}
    
        insert_into_audit(audit_data)
    except Exception:
        logging.exception("issue in the query formation")
    return jsonify(return_json_data)


##### From here the below code is responsible to generate the extraction Report
def segregrate_fields(formated_fields,edited_fields,extraction_fields,cols_to_igonre):
    for case_data in formated_fields:
        edited_fields_cnt=0
        extracted_cnt=0
        unused_field_cnt=0
        for k,v in case_data.items():
            if k in edited_fields and k not in cols_to_igonre:
                edited_fields_cnt=edited_fields_cnt+1
                case_data[k]=f"M:{v}"
            elif k in extraction_fields and str(v).strip()!= "" and v is not None and k not in cols_to_igonre:
                extracted_cnt=extracted_cnt+1
                case_data[k]=f"E:{v}"
            elif k not in cols_to_igonre:
                unused_field_cnt=unused_field_cnt+1
            else:
                pass
        case_data['Number of fields']=97
        case_data['Manual']=edited_fields_cnt
        case_data['Extraction']=extracted_cnt
        case_data['Not used fields']=unused_field_cnt
        case_data['Extraction Percentage']=100
    return formated_fields

def parse_fields(val):
    fields={}
    try:
        tab_view=val.get('tab_view',{})
        # print(f"tab VIEW IS {tab_view}")
        if tab_view:
            row_data=tab_view['rowData']
            for data in row_data:
                # print(f"data got is {data}")
                key=data.get('fieldName')
                value=data.get('value')
                if not isinstance(value, dict):
                    fields[key] = value
    except Exception as e:
        logging.exception(f"## Exception occurred while parsing fields ... {e}")
        pass
    return fields

def format_fields(new_query_df,edited_fields):
    print(f" ## Running format fields ")
    casewise_formatted_fields=[]
    try:
        for case_id in new_query_df:
            formatted_fields={}
            for key,value in case_id.items():
                try:
                    value=json.loads(value)
                except:
                    pass
                if isinstance(value,dict):
                    parsed_fields=parse_fields(value)
                    formatted_fields.update(parsed_fields)
                else:
                    formatted_fields[key]=value 
            casewise_formatted_fields.append(formatted_fields)
    except Exception as e:
        logging.exception(f"## Exception occurred while format fields .. {e}") 
        pass  
    return casewise_formatted_fields

def generate_summary_data(report_data_df):
    summary_data={}
    summary_data['Number_of_Cases']=report_data_df['case_id'].count()
    summary_data['Date_of_inputting']='10-05-2025'
    summary_data['Status']='Approved'
    summary_data['Number_of_fields']=97
    summary_data['Extraction']=round(report_data_df['Extraction'].mean(),1)
    summary_data['Manual']=round(report_data_df['Manual'].mean(),1)
    summary_data['Not_used_fields']=round(report_data_df['Not used fields'].mean(),1)
    summary_data['Extraction_Percentage']=  round(report_data_df['Extraction Percentage'].mean(),2)

    summary_data_df=pd.DataFrame([summary_data])				
    return summary_data_df
 
    
def merge_all_jsons(row):
    try:
        d1 = json.loads(clob_to_string(row['NOT_EXTRACTED_FIELDS'])) if row['NOT_EXTRACTED_FIELDS'] else {}
        d2 = json.loads(clob_to_string(row['FIELDS_EDITED'])) if row['FIELDS_EDITED'] else {}
        d3 = json.loads(clob_to_string(row['FIELDS_EXTRACTED'])) if row['FIELDS_EXTRACTED'] else {}
        return {**d1, **d2, **d3}  # Merge dicts (later keys overwrite earlier ones if duplicated)
    except Exception as e:
        print(f"Error parsing row {row.name}: {e}")
        return {}

def clob_to_string(clob):
    if clob:
        # If it's a CLOB object, read it as string
        return clob.read() if hasattr(clob, 'read') else str(clob)
    return "{}"

def normalize(col_name):
    return col_name.strip().lower().replace(" ", " ").replace("\xa0", " ")

# Determines which columns have any non-empty & non-NaN values
def has_real_values(col):
    return ((~col.isna()) & (col != "")).any()

@app.route('/extraction_report', methods=['POST', 'GET'])
def extraction_report():
    logging.info('Calling Extraction Report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except Exception:
        logging.warning("#Failed to start ram and time calc#")
    
    try:
        # Validate the incoming request
        if not request.is_json:
            return jsonify({'flag': False, 'message': 'Invalid Request - JSON expected'})
            
        data = request.json
        if not data or 'ui_data' not in data or not data.get('start_date') or not data.get('end_date'):
            return jsonify({'flag': False, 'message': 'Invalid Request - Missing required fields'})
        
        tenant_id = data['ui_data']['tenant_id']
        if not tenant_id:
            return jsonify({'flag': False, 'message': 'Invalid Request - Missing tenant_id'})
            
        # Tracing setup
        trace_id = generate_random_64bit_string()
        attr = ZipkinAttrs(
            trace_id=generate_random_64bit_string(),
            span_id=generate_random_64bit_string(),
            parent_span_id=None,
            flags=None,
            is_sampled=False,
            tenant_id=tenant_id
        )

        with zipkin_span(
                service_name='reports_api',
                zipkin_attrs=attr,
                span_name='extraction_report',
                transport_handler=http_transport,
                sample_rate=0.5
        ):
            # Get report parameters
            reports_name = 'Extraction Report'
            logging.info(f"reports_name: {reports_name}")
            
            db_config['tenant_id'] = tenant_id
            queues_db = DB("queues", **db_config)
            extraction_db = DB("extraction", **db_config)

            start_date = data['start_date']
            end_date = data['end_date']
            ist = pytz.timezone("Asia/Calcutta")
            timestamp = datetime.now(ist)        
            timestamp1 = timestamp.strftime('%d-%m-%Y %H:%M:%S')
            timestamp = str(timestamp)[:-13]
            
            # Query to fetch case data
            query = f"""SELECT 
                o.PARTY_ID as "PARTY ID", 
                pq.hub_code as "HUB Code",
                pq.branch_code as "Branch code",
                pq.branch_name as "Branch Location",
                o.CUSTOMER_NAME as "Customer Name",
                pm.RELATION_MGR_EMP_CODE as "RM Code",
                o.RELATION_MGRNAME as "RM Name",
                pq.FILE_NAME as "File Name", 
                CAST(FROM_TZ(CAST(pq.CREATED_DATE AS TIMESTAMP), 'UTC') AT TIME ZONE 'Asia/Kolkata' AS TIMESTAMP) AS "Case Creation date", 
                pq.REGION as "WBO Region",
                o.RM_STATE as "RM State",
                o.CASE_ID as "Case ID", 
                pq.STATUS as "STATUS",
                pq.LAST_UPDATED_BY as "Maker ID",  
                pq.approved_date  as "Date of inputting",
                o.DATE_STAT as "Statement for the Month", 
                o.STOCK_DOC_MONTH as "Stock Doc Month", 
                o.STOCK_DOC_YEAR as "Stock Doc Year", 
                o.DUE_DATE as "Due Date", 
                o.DP_TO_BE_CALCULATED as "DP to be calculated manually",
                o.DP_ALLOCATED_BY_CUSTOMER as "DP Allocated by Customer", 
                o.BANK_SHARE as "Bank share",
                o.BANK_SHARE_PERCENTAGE as "Bank Share Percentage",
                o.CUT_OFF_PERCENTAGE as "Cut off Percentage", 
                o.R1_RATING as "RR 1 Rating", 
                o.R1_PLUS_RATING as "RR 1+ Rating", 
                o.R2_RATING as "RR 2 Rating",
                o.R2_PLUS_SCRIPTS as "RR 2+ Rating",
                o.R3_SCRIPTS as "RR 3 Rating",
                o.REJECTED_COMMENTS as "Rejected Comments",
                o.REMARKS as "REMARKS",
                o.UNHOLD_COMMENTS as "Unhold Justification",
                o.COMMENTS as "Hold Justification",
                fa.number_of_fields as "Total Fields",
                fa.extraction as "Extraction",
                fa.manual as "Manual",
                fa.extraction_percentage as "Extraction Percentage",
                fa.manual_percentage as "Manual Percentage",
                fa.NOT_EXTRACTED_FIELDS,
                fa.FIELDS_EDITED,
                fa.FIELDS_EXTRACTED,
                o.CASE_COUNT as "Count of ingestion",
                o.STOCK_SEC_REFERENCE_ID as "Doc ID",
                pq.PROCESS_CNT as "Processing Count"
            FROM hdfc_queues.process_queue pq 
            INNER JOIN hdfc_queues.queue_list ql 
                ON pq.case_id = ql.case_id 
            INNER JOIN hdfc_extraction.ocr o 
                ON pq.case_id = o.case_id
            INNER JOIN hdfc_queues.field_accuracy fa 
                ON pq.case_id = fa.case_id
            LEFT JOIN hdfc_extraction.party_master pm 
                ON o.PARTY_ID = CAST(pm.PARTY_ID AS VARCHAR2(4000))
            WHERE pq.created_date BETWEEN TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')
            AND ql.queue != 'case_creation'
            """
            query_df = queues_db.execute_(query)
            #logging.info(f'query_df----------{query_df}')
            
            # Extract merged fields (dynamic fields)
            query_df['merged_fields'] = query_df.apply(merge_all_jsons, axis=1)
            merged_expanded_df = pd.json_normalize(query_df['merged_fields'])
            query_df = query_df.drop(columns=['merged_fields', 'NOT_EXTRACTED_FIELDS', 'FIELDS_EDITED', 'FIELDS_EXTRACTED','not_extracted_fields','fields_edited','fields_extracted','status','remarks'])

            # Add dynamic fields to DataFrame
            dynamic_columns = list(merged_expanded_df.columns)
            for col in dynamic_columns:
                if col not in query_df.columns:
                    query_df[col] = merged_expanded_df[col]
            

            desired_column_order = ["PARTY ID", "Customer Name", "RM Code", "RM Name", "Doc ID", "File Name", "Case Creation date", "Branch code", "Branch Location","HUB Code", "WBO Region", "RM State", "Count of ingestion", "Processing Count", "Maker ID", "Date of inputting", "Case ID", "STATUS","date_stat", "Statement for the Month","Stock Doc Month", "Stock Doc Year", "Due Date", "Bank share", "Raw Materials", "Work in Process", "Finished Goods", "Total Stock","Stores and Spares", "Stock in Transit", "Bullion stock", "Obselete stock", "Standard Gold", "Pledge stock", "Consumable and Spares","Goods in Transit", "PCFC Stock", "Inventory Financed stock", "Domestic Stock", "Export Stock", "Indegenious Stock", "Receivables","Total Debtors", "Book debts upto 7 days", "Debtors <30 days", "Debtors <60 days", "Debtors <90 days", "Debtors <120 days", "Debtors <150 days","Debtors <180 days", "Debtors >180 days", "Debtors Exports", "Debtors Domestic", "Debtors of Group Company", "Unbilled Debtors", "Domestic receivables","Export receivables", "Gst Receivables", "Sales", "Purchases", "Customer Duty Refundable", "Duty Drawback", "Government Receivables", "Unbilled Revenue","Retention Money", "Unbilled Receivables", "Buyer Credit Outstanding", "Debtors PC", "Total Creditors", "Creditors Domestic", "Creditors Exports","Creditors PC", "Unpaid Stocks", "GSS(gold saving scheme)", "DALC", "LC Creditors", "Stock under LC", "Margin Money","Fixed Deposit", "Trade Creditors", "Group Company Creditors", "Advances from Debtors/Customers", "Advances paid to suppliers", "Under Deposits/Bhishis Schemes","Channel Finance", "Advance BG outstanding", "Purchase Invoice Discounting", "Sales Invoice Discounting", "Term Loan", "Outstanding Amount", "Assest Cover Ratio","HDFC BANK_DP Allocated", "Total utilisations", "Cut of Percentage", "RR 1 Rating", "RR 1+ Rating", "RR 2 Rating", "RR 2+ Rating", "RR 3 Rating", "DP to be calculated manually","DP Allocated by Customer", "Bank Share Percentage", "DP As Per Lead Bank"] 
            # Add component fields (stock_fields, debtor_fields, creditor_fields, advance_fields)

            # Query to get the component fields for stock, debtors, creditors, advances
            component_query = """
                SELECT DISTINCT CAST(cm.component_name AS VARCHAR2(4000)) AS component_name, CAST(cm.component_category AS VARCHAR2(4000)) AS component_category
                FROM COMPONENT_MASTER cm
                  WHERE CAST(cm.status AS VARCHAR2(4000)) = 'ACTIVE'
                AND (CAST(cm.component_category AS VARCHAR2(4000)) = 'Stock'
                     OR CAST(cm.component_category AS VARCHAR2(4000)) = 'Debtors'
                     OR CAST(cm.component_category AS VARCHAR2(4000)) = 'Creditors'
                     OR CAST(cm.component_category AS VARCHAR2(4000)) = 'Advances')
            """
            component_df = extraction_db.execute_(component_query)  # Assuming `execute_` is a method to fetch the query result
            
            # Extracting component fields by category
            stock_fields = component_df[component_df['component_category'] == 'Stock']['component_name'].tolist()
            debtor_fields = component_df[component_df['component_category'] == 'Debtors']['component_name'].tolist()
            creditor_fields = component_df[component_df['component_category'] == 'Creditors']['component_name'].tolist()
            advance_fields = component_df[component_df['component_category'] == 'Advances']['component_name'].tolist()

	    
            try:
                query_df["Statement for the Month"] = query_df["date_stat"]
            except Exception as e:
                logging.info(f" the statement for the month is not changed as {e}")

            try:
                if "date_stat" in query_df.columns:
                    query_df = query_df.drop(columns=["date_stat"])
            except Exception as E:
                logging.info(f" date_stat is not dropped as {E}")

            # Add non-empty component fields only
            #for col in stock_fields + debtor_fields + creditor_fields + advance_fields:
            #    if col in query_df.columns and query_df[col].notna().any():  # Check if the column has non-null values
            #        if col not in query_df.columns:
            #            continue

# Keep only component fields that exist AND have non-null values
            #stock_fields = [col for col in stock_fields if col in query_df.columns and query_df[col].notna().any()]
            stock_fields_updated = [col for col in stock_fields if col in query_df.columns and has_real_values(query_df[col])]
            debtor_fields_updated = [col for col in debtor_fields if col in query_df.columns and has_real_values(query_df[col])]
            creditor_fields_updated = [col for col in creditor_fields if col in query_df.columns and has_real_values(query_df[col])]
            advance_fields_updated = [col for col in advance_fields if col in query_df.columns and has_real_values(query_df[col])]

            # Define the desired column order
            #col_order =  desired_column_order + stock_fields + debtor_fields + creditor_fields + advance_fields 
            #valid_columns = [col for col in col_order if col in query_df.columns]
            last_cols = [ "REMARKS","Rejected Comments", "Hold Justification", "Unhold Justification","Total Fields", "Extraction", "Manual", "Extraction Percentage", "Manual Percentage"]
            desired_column_order= [c for c in desired_column_order if c not in last_cols]
            component_fields=stock_fields + debtor_fields + creditor_fields + advance_fields
            col_order =  desired_column_order + stock_fields_updated + debtor_fields_updated + creditor_fields_updated + advance_fields_updated 
            valid_columns = [col for col in col_order if col in query_df.columns]
            #extra_columns = [col for col in query_df.columns if col not in col_order and not in component_fields]
            #valid_columns.extend(extra_columns)
            #final_columns=valid_columns+last_cols
            extra_columns = [col for col in query_df.columns if col not in col_order and col not in last_cols and col not in component_fields]
            logging.info(f"the extra columns are -----{extra_columns}")
            final_columns=valid_columns+extra_columns
            final_columns_order=final_columns+last_cols

            # Reorganize the dataframe based on the desired order
            final_df = query_df[final_columns_order]
            final_df = final_df.loc[:, ~final_df.columns.duplicated()]
            final_df = final_df.loc[
                :,
                ~final_df.columns.str.startswith("HDFC BANK_DP Allocated")
            ]




            # Save to Excel file
            parsed_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")

            filename = f'Extraction_report--{formatted_date}.xlsx'
            output_file = f'/var/www/reports_api/reports/{filename}'

            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name="Extraction Report", index=False)

            end_time = tt()
            time_consumed = str(round(end_time - start_time, 3))
            logging.info(f"Time Consumed for Extraction report is {time_consumed}")

            return_json_data = {}
            return_json_data['message'] = 'Successfully generated the report'
            return_json_data['flag'] = True
            return_json_data['excel_flag'] = 1
            return_json_data['data'] = {'row_data': final_df.to_dict(orient='records')}
            data['report_data'] = return_json_data
            return jsonify(return_json_data)

    except Exception as e:
        logging.exception(f'Something went wrong exporting data: {e}')
        return_json_data = {}
        return_json_data['flag'] = False
        return_json_data['message'] = 'Failed!!'
        return jsonify(return_json_data)